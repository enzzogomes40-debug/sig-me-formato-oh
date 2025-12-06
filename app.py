from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file, send_from_directory
import datetime
import logging
import mysql.connector
from mysql.connector import Error
import os
import csv
import io
import json
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

# =============================================================================
# CONFIGURAÇÕES PARA PYTHONANYWHERE
# =============================================================================
import sys

# Configuração para PythonAnywhere
if 'PYTHONANYWHERE_DOMAIN' in os.environ:
    # Desativa debug no ambiente de produção
    DEBUG = False
    # Configurações específicas para PythonAnywhere
    MYSQL_HOST = 'enzzodril.mysql.pythonanywhere-services.com'
    MYSQL_DATABASE = 'enzzodril$default'
    MYSQL_USER = 'enzzodril'
    MYSQL_PASSWORD = '123formato'
else:
    # Configurações para desenvolvimento local
    DEBUG = True
    MYSQL_HOST = 'localhost'
    MYSQL_DATABASE = 'sig_me_db'
    MYSQL_USER = 'root'
    MYSQL_PASSWORD = '123456'

# CONFIGURAR LOGGING
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
app = Flask(__name__)
app.secret_key = 'sig-me-chave-secreta-2024'

# Configuração do app para PythonAnywhere
if 'PYTHONANYWHERE_DOMAIN' in os.environ:
    app.config.update(
        DEBUG=False,
        SESSION_COOKIE_SECURE=True,
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE='Lax',
        PERMANENT_SESSION_LIFETIME=3600
    )

# CORES DO SITE FORMATO OH - EM FORMATO ARGB PARA EXCEL
CORES_FORMATO = {
    'primaria': '#1a365d',      # Azul escuro
    'secundaria': '#2d3748',    # Cinza azulado
    'destaque': '#e53e3e',      # Vermelho
    'sucesso': '#38a169',       # Verde
    'alerta': '#dd6b20',        # Laranja
    'info': '#3182ce',          # Azul
    'claro': '#f7fafc',         # Cinza muito claro
    'branco': '#ffffff',
    'texto': '#2d3748',
    'texto_claro': '#718096'
}

# Converter cores hex para ARGB (adicionar 'FF' no início)
def hex_to_argb(hex_color):
    """Converte cor hex para formato ARGB do Excel"""
    hex_color = hex_color.lstrip('#')
    return 'FF' + hex_color

# CORES EM FORMATO ARGB PARA USO NO EXCEL
CORES_FORMATO_ARGB = {
    'primaria': hex_to_argb(CORES_FORMATO['primaria']),
    'secundaria': hex_to_argb(CORES_FORMATO['secundaria']),
    'destaque': hex_to_argb(CORES_FORMATO['destaque']),
    'sucesso': hex_to_argb(CORES_FORMATO['sucesso']),
    'alerta': hex_to_argb(CORES_FORMATO['alerta']),
    'info': hex_to_argb(CORES_FORMATO['info']),
    'claro': hex_to_argb(CORES_FORMATO['claro']),
    'branco': hex_to_argb(CORES_FORMATO['branco']),
}

# =============================================================================
# CONFIGURAÇÃO PARA ARQUIVOS PDF
# =============================================================================
# Diretório para uploads de PDF (criar se não existir)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
PDF_FOLDER = os.path.join(UPLOAD_FOLDER, 'contratos')

# Criar diretórios se não existirem
os.makedirs(PDF_FOLDER, exist_ok=True)

# Configurar o Flask para servir arquivos estáticos dos diretórios de upload
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PDF_FOLDER'] = PDF_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# =============================================================================
# MÓDULO: DATABASE - CONEXÃO MYSQL REAL
# =============================================================================
class DatabaseConnection:
    def __init__(self):
        self.host = MYSQL_HOST
        self.database = MYSQL_DATABASE
        self.user = MYSQL_USER
        self.password = MYSQL_PASSWORD
        self.connection = None

    def connect(self, create_database=False):
        """Conecta ao MySQL real"""
        try:
            if create_database:
                self.connection = mysql.connector.connect(
                    host=self.host,
                    user=self.user,
                    password=self.password
                )
            else:
                self.connection = mysql.connector.connect(
                    host=self.host,
                    database=self.database,
                    user=self.user,
                    password=self.password
                )

            if self.connection.is_connected():
                logging.info("✅ Conectado ao MySQL com sucesso!")
                return True
        except Error as e:
            logging.error(f"❌ Erro de conexão MySQL: {e}")
            return False

    def execute_query(self, query, params=None):
        """Executa queries no MySQL real"""
        if not self.connect():
            return None

        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(query, params or ())
            result = cursor.fetchall()
            cursor.close()

            # Converter objetos Decimal para float
            for row in result:
                for key, value in row.items():
                    if isinstance(value, Decimal):
                        row[key] = float(value)

            return result
        except Error as e:
            logging.error(f"❌ Erro na query: {e}")
            return None
        finally:
            if self.connection and self.connection.is_connected():
                self.connection.close()

    def execute_update(self, query, params=None):
        """Executa updates no MySQL real"""
        if not self.connect():
            return False

        try:
            cursor = self.connection.cursor()
            cursor.execute(query, params or ())
            self.connection.commit()
            cursor.close()
            logging.info(f"✅ Update executado: {query}")
            return True
        except Error as e:
            logging.error(f"❌ Erro no update: {e}")
            return False
        finally:
            if self.connection and self.connection.is_connected():
                self.connection.close()

# =============================================================================
# INICIALIZAR DATABASE E CRIAR TABELAS SE NÃO EXISTIREM
# =============================================================================
def inicializar_database():
    """Inicializa o banco de dados e cria tabelas se não existirem"""
    try:
        db = DatabaseConnection()

        # Conectar sem database específico para criar se não existir
        conn = mysql.connector.connect(
            host=db.host,
            user=db.user,
            password=db.password
        )
        cursor = conn.cursor()

        # Criar database se não existir
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {db.database}")
        cursor.execute(f"USE {db.database}")

        # === CRIAR TABELA DE CLIENTES ===
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                Nome_Fantasia VARCHAR(255) NOT NULL,
                Nome_Razao_Social VARCHAR(255),
                CNPJ_CPF VARCHAR(20),
                Telefone VARCHAR(20),
                Email VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Criar tabela de fornecedores se não existir
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS fornecedores (
                id INT AUTO_INCREMENT PRIMARY KEY,
                razao_social VARCHAR(255) NOT NULL,
                cnpj VARCHAR(20),
                telefone VARCHAR(20),
                email VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Criar tabela de contratos
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS contratos (
                id INT AUTO_INCREMENT PRIMARY KEY,
                tipo ENUM('cliente', 'fornecedor') NOT NULL,
                cliente_id INT,
                fornecedor_id INT,
                descricao TEXT NOT NULL,
                valor_mensal DECIMAL(10,2) NOT NULL,
                data_inicio DATE NOT NULL,
                data_fim DATE,
                status ENUM('ativo', 'inativo', 'vencido') DEFAULT 'ativo',
                observacoes TEXT,
                arquivo_pdf VARCHAR(255),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Criar tabela de placas (inventário)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS placas (
                id INT AUTO_INCREMENT PRIMARY KEY,
                Codigo_Ativo VARCHAR(100) NOT NULL,
                Endereco TEXT NOT NULL,
                Regiao VARCHAR(100) NOT NULL,
                Tipo_Placa VARCHAR(100) NOT NULL,
                Status_Atual ENUM('disponível', 'locado', 'reservado', 'manutenção') DEFAULT 'disponível',
                Cliente_Locacao VARCHAR(255),
                Valor_Mensal DECIMAL(10,2) DEFAULT 0,
                Data_Cadastro DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Criar tabela de transações financeiras
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS transacoes_financeiras (
                id INT AUTO_INCREMENT PRIMARY KEY,
                tipo ENUM('receita', 'despesa') NOT NULL,
                descricao VARCHAR(255) NOT NULL,
                valor DECIMAL(10,2) NOT NULL,
                categoria VARCHAR(100) NOT NULL,
                data_transacao DATE NOT NULL,
                data_vencimento DATE,
                status ENUM('pendente', 'pago', 'atrasado') DEFAULT 'pendente',
                cliente_id INT,
                fornecedor_id INT,
                observacoes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Inserir alguns clientes de exemplo se a tabela estiver vazia
        cursor.execute("SELECT COUNT(*) as count FROM clientes")
        if cursor.fetchone()[0] == 0:
            clientes_exemplo = [
                ('Concessionária XYZ', 'Concessionária XYZ Ltda', '12.345.678/0001-90', '(11) 3333-4444', 'contato@xyz.com.br'),
                ('Rede de Farmácias', 'Rede de Farmácias SA', '98.765.432/0001-10', '(11) 5555-6666', 'vendas@rededefarmacias.com.br'),
                ('Banco Nacional', 'Banco Nacional SA', '55.444.333/0001-20', '(11) 7777-8888', 'contratos@bancacional.com.br')
            ]

            for cliente in clientes_exemplo:
                cursor.execute(
                    "INSERT INTO clientes (Nome_Fantasia, Nome_Razao_Social, CNPJ_CPF, Telefone, Email) VALUES (%s, %s, %s, %s, %s)",
                    cliente
                )

        # Inserir alguns fornecedores de exemplo se a tabela estiver vazia
        cursor.execute("SELECT COUNT(*) as count FROM fornecedores")
        if cursor.fetchone()[0] == 0:
            fornecedores_exemplo = [
                ('Construtora Silva Ltda', '12.345.678/0001-90', '(11) 9999-8888', 'contato@silvaconstrucoes.com.br'),
                ('Iluminação Publica SP', '98.765.432/0001-10', '(11) 7777-6666', 'contato@iluminacaosp.com.br'),
                ('Manutenção Express', '55.444.333/0001-20', '(11) 5555-4444', 'servicos@manutencaoexpress.com.br')
            ]

            for fornecedor in fornecedores_exemplo:
                cursor.execute(
                    "INSERT INTO fornecedores (razao_social, cnpj, telefone, email) VALUES (%s, %s, %s, %s)",
                    fornecedor
                )

        # Inserir alguns contratos de exemplo se a tabela estiver vazia
        cursor.execute("SELECT COUNT(*) as count FROM contratos")
        if cursor.fetchone()[0] == 0:
            contratos_exemplo = [
                ('cliente', 1, None, 'Contrato de Locação - Concessionária XYZ', 3500.00, '2024-01-01', '2024-12-31', 'ativo', 'Contrato anual renovável', 'contrato_xyz.pdf'),
                ('cliente', 2, None, 'Locação Especial - Rede Farmácias', 3200.00, '2024-02-01', '2024-12-31', 'ativo', 'Campanha temporária', 'contrato_farmacias.pdf'),
                ('cliente', 3, None, 'Contrato Banco Nacional', 2800.00, '2024-03-01', '2024-12-31', 'ativo', 'Contrato corporativo', 'contrato_banco.pdf')
            ]

            for contrato in contratos_exemplo:
                cursor.execute(
                    "INSERT INTO contratos (tipo, cliente_id, fornecedor_id, descricao, valor_mensal, data_inicio, data_fim, status, observacoes, arquivo_pdf) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    contrato
                )

        # Inserir algumas placas de exemplo se a tabela estiver vazia
        cursor.execute("SELECT COUNT(*) as count FROM placas")
        if cursor.fetchone()[0] == 0:
            placas_exemplo = [
                # São Roque - 8 placas
                ('43', 'Rod Raposo Tavares km S8 - Semi. Centro de São Roque', 'São Roque', 'Outdoor', 'locado', 'Concessionária XYZ', 3500.00, '2024-01-15'),
                ('44-2', 'Rod Raposo Tavares km S8 - Entrada de São Roque', 'São Roque', 'Outdoor', 'locado', 'Rede de Farmácias', 3200.00, '2024-01-15'),
                ('141', 'Av. Principal - Centro', 'São Roque', 'Outdoor', 'locado', 'Banco Nacional', 2800.00, '2024-02-01'),
                ('142', 'Av. Getúlio Vargas, 500 - Centro', 'São Roque', 'Front Light', 'locado', 'Supermercado Preço Baixo', 2200.00, '2024-02-10'),
                ('143', 'Rua das Flores, 123 - Jardim Europa', 'São Roque', 'Back Light', 'reservado', 'Construtora Progresso', 1800.00, '2024-02-15'),
                ('144', 'Rodovia Prefeito Quintino de Lima', 'São Roque', 'Outdoor', 'disponível', None, 2500.00, '2024-03-01'),
                ('145', 'Praça da Matriz - Centro Histórico', 'São Roque', 'Led', 'manutenção', None, 0.00, '2024-03-05'),
                ('146', 'Av. João Pessoa, 789 - Vila Nova', 'São Roque', 'Front Light', 'disponível', None, 1900.00, '2024-03-10'),

                # Mairinque - 6 placas
                ('201', 'Rod. Bunjiro Nakao, km 45', 'Mairinque', 'Outdoor', 'locado', 'Posto Shell', 3100.00, '2024-01-20'),
                ('202', 'Av. São Paulo, 456 - Centro', 'Mairinque', 'Front Light', 'locado', 'Farmácia Popular', 2100.00, '2024-02-05'),
                ('203', 'Rua Amazonas, 321 - Vila Maria', 'Mairinque', 'Back Light', 'disponível', None, 1700.00, '2024-02-12'),
                ('204', 'Entrada da Cidade - BR-376', 'Mairinque', 'Outdoor', 'reservado', 'Concessionária ABC', 2900.00, '2024-02-20'),
                ('205', 'Praça Central - Centro', 'Mairinque', 'Led', 'disponível', None, 2300.00, '2024-03-01'),
                ('206', 'Av. Industrial, 1000 - Distrito Industrial', 'Mairinque', 'Front Light', 'locado', 'Indústria Metalúrgica', 2000.00, '2024-03-08'),

                # Ibuna - 5 placas
                ('301', 'Estrada Municipal Ibuna-São Roque', 'Ibuna', 'Outdoor', 'locado', 'Laticínios Serra', 2700.00, '2024-01-25'),
                ('302', 'Centro Comercial de Ibuna', 'Ibuna', 'Front Light', 'disponível', None, 1600.00, '2024-02-08'),
                ('303', 'Entrada do Bairro Rural', 'Ibuna', 'Back Light', 'disponível', None, 1400.00, '2024-02-14'),
                ('304', 'Praça da Igreja Matriz', 'Ibuna', 'Outdoor', 'locado', 'Mercado Central', 2400.00, '2024-02-25'),
                ('305', 'Rod. Vicinal Principal', 'Ibuna', 'Front Light', 'manutenção', None, 0.00, '2024-03-03'),

                # Araçá - 6 placas
                ('401', 'Rua Principal - Centro do Araçá', 'Araçá', 'Outdoor', 'locado', 'Auto Peças Araçá', 2600.00, '2024-01-30'),
                ('402', 'Av. Comercial, 200', 'Araçá', 'Front Light', 'disponível', None, 1800.00, '2024-02-12'),
                ('403', 'Praça do Mercado Municipal', 'Araçá', 'Back Light', 'reservado', 'Panificadora Doce Pão', 1500.00, '2024-02-18'),
                ('404', 'Entrada da Cidade - SP-250', 'Araçá', 'Outdoor', 'locado', 'Transportadora Rápida', 2800.00, '2024-02-28'),
                ('405', 'Rua da Escola Municipal', 'Araçá', 'Led', 'disponível', None, 2100.00, '2024-03-05'),
                ('406', 'Bairro Novo Horizonte', 'Araçá', 'Front Light', 'disponível', None, 1700.00, '2024-03-12'),

                # Piedade - 7 placas
                ('501', 'Rod. Padre Anchieta, km 78', 'Piedade', 'Outdoor', 'locado', 'Hotel Serra Verde', 3300.00, '2024-02-01'),
                ('502', 'Centro Histórico de Piedade', 'Piedade', 'Front Light', 'locado', 'Restaurante Sabor Caseiro', 2200.00, '2024-02-10'),
                ('503', 'Av. Beira Rio, 150', 'Piedade', 'Back Light', 'disponível', None, 1600.00, '2024-02-15'),
                ('504', 'Entrada do Parque Municipal', 'Piedade', 'Outdoor', 'reservado', 'Agência de Turismo', 3000.00, '2024-02-22'),
                ('505', 'Praça do Santuário', 'Piedade', 'Led', 'locado', 'Loja de Artigos Religiosos', 1900.00, '2024-03-01'),
                ('506', 'Rua Comercial, 345', 'Piedade', 'Front Light', 'disponível', None, 2000.00, '2024-03-08'),
                ('507', 'Bairro dos Pinheiros', 'Piedade', 'Back Light', 'manutenção', None, 0.00, '2024-03-15'),

                # Alumínio - 5 placas
                ('601', 'Rod. dos Metalúrgicos, km 12', 'Alumínio', 'Outdoor', 'locado', 'Indústria de Alumínio', 3200.00, '2024-02-05'),
                ('602', 'Centro de Alumínio', 'Alumínio', 'Front Light', 'disponível', None, 2100.00, '2024-02-12'),
                ('603', 'Av. Industrial, 500', 'Alumínio', 'Back Light', 'locado', 'Metalúrgica Forte', 2300.00, '2024-02-20'),
                ('604', 'Entrada do Complexo Industrial', 'Alumínio', 'Outdoor', 'reservado', 'Cooperativa de Trabalho', 2900.00, '2024-02-28'),
                ('605', 'Praça dos Trabalhadores', 'Alumínio', 'Led', 'disponível', None, 2400.00, '2024-03-06')
            ]

            for placa in placas_exemplo:
                cursor.execute(
                    "INSERT INTO placas (Codigo_Ativo, Endereco, Regiao, Tipo_Placa, Status_Atual, Cliente_Locacao, Valor_Mensal, Data_Cadastro) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                    placa
                )

        # Inserir algumas transações financeiras de exemplo
        cursor.execute("SELECT COUNT(*) as count FROM transacoes_financeiras")
        if cursor.fetchone()[0] == 0:
            transacoes_exemplo = [
                # Receitas
                ('receita', 'Locação Placa 43 - Concessionária XYZ', 3500.00, 'Locação Placas', '2024-03-01', '2024-03-01', 'pago', 1, None, 'Pagamento recebido'),
                ('receita', 'Locação Placa 44-2 - Rede Farmácias', 3200.00, 'Locação Placas', '2024-03-05', '2024-03-05', 'pago', 2, None, 'Pagamento recebido'),
                ('receita', 'Locação Placa 141 - Banco Nacional', 2800.00, 'Locação Placas', '2024-03-10', '2024-03-10', 'pago', 3, None, 'Pagamento recebido'),
                ('receita', 'Locação Placa 142 - Supermercado', 2200.00, 'Locação Placas', '2024-03-15', '2024-03-15', 'pendente', None, None, 'Aguardando pagamento'),

                # Despesas
                ('despesa', 'Manutenção Placa 145', 450.00, 'Manutenção', '2024-03-02', '2024-03-02', 'pago', None, 3, 'Manutenção preventiva'),
                ('despesa', 'Aluguel Escritório', 1200.00, 'Despesas Operacionais', '2024-03-05', '2024-03-05', 'pago', None, None, 'Aluguel mensal'),
                ('despesa', 'Energia Elétrica', 380.00, 'Despesas Operacionais', '2024-03-08', '2024-03-15', 'pendente', None, None, 'Conta a vencer'),
                ('despesa', 'Material de Limpeza', 150.00, 'Despesas Operacionais', '2024-03-10', '2024-03-10', 'pago', None, None, 'Material de escritório')
            ]

            for transacao in transacoes_exemplo:
                cursor.execute(
                    "INSERT INTO transacoes_financeiras (tipo, descricao, valor, categoria, data_transacao, data_vencimento, status, cliente_id, fornecedor_id, observacoes) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    transacao
                )

        conn.commit()
        cursor.close()
        conn.close()

        logging.info("✅ Banco de dados inicializado com sucesso!")
        return True

    except Error as e:
        logging.error(f"❌ Erro ao inicializar banco de dados: {e}")
        return False

# =============================================================================
# MÓDULO: RELATÓRIOS
# =============================================================================
class RelatorioManager:
    def __init__(self, db_connection):
        self.db = db_connection

    def gerar_relatorio_ocupacao(self):
        """Gera relatório de ocupação detalhado"""
        try:
            query = """
                SELECT Regiao, Status_Atual, COUNT(*) as quantidade,
                       SUM(Valor_Mensal) as faturamento
                FROM placas
                GROUP BY Regiao, Status_Atual
                ORDER BY Regiao, Status_Atual
            """
            dados = self.db.execute_query(query) or []

            # Agrupar por região
            relatorio = {}
            for item in dados:
                regiao = item['Regiao']
                if regiao not in relatorio:
                    relatorio[regiao] = []
                relatorio[regiao].append(item)

            return relatorio

        except Exception as e:
            logging.error(f"❌ Erro ao gerar relatório ocupação: {e}")
            return {}

    def gerar_relatorio_financeiro(self):
        """Gera relatório financeiro detalhado"""
        try:
            # Faturamento por região
            faturamento_regiao = self.db.execute_query("""
                SELECT Regiao, SUM(Valor_Mensal) as faturamento
                FROM placas
                WHERE Status_Atual IN ('locado', 'reservado')
                GROUP BY Regiao
                ORDER BY faturamento DESC
            """) or []

            # Faturamento por tipo de placa
            faturamento_tipo = self.db.execute_query("""
                SELECT Tipo_Placa, SUM(Valor_Mensal) as faturamento
                FROM placas
                WHERE Status_Atual IN ('locado', 'reservado')
                GROUP BY Tipo_Placa
                ORDER BY faturamento DESC
            """) or []

            # Top clientes
            top_clientes = self.db.execute_query("""
                SELECT Cliente_Locacao, SUM(Valor_Mensal) as faturamento
                FROM placas
                WHERE Status_Atual IN ('locado', 'reservado') AND Cliente_Locacao IS NOT NULL
                GROUP BY Cliente_Locacao
                ORDER BY faturamento DESC
                LIMIT 10
            """) or []

            return {
                'faturamento_regiao': faturamento_regiao,
                'faturamento_tipo': faturamento_tipo,
                'top_clientes': top_clientes,
                'faturamento_total': sum(item['faturamento'] or 0 for item in faturamento_regiao)
            }

        except Exception as e:
            logging.error(f"❌ Erro ao gerar relatório financeiro: {e}")
            return {}

    def gerar_relatorio_regiao(self):
        """Gera relatório por região"""
        try:
            query = """
                SELECT Regiao,
                       COUNT(*) as total_placas,
                       SUM(CASE WHEN Status_Atual = 'disponível' THEN 1 ELSE 0 END) as disponiveis,
                       SUM(CASE WHEN Status_Atual = 'locado' THEN 1 ELSE 0 END) as locadas,
                       SUM(CASE WHEN Status_Atual = 'reservado' THEN 1 ELSE 0 END) as reservadas,
                       SUM(CASE WHEN Status_Atual = 'manutenção' THEN 1 ELSE 0 END) as manutencao,
                       SUM(Valor_Mensal) as faturamento_total
                FROM placas
                GROUP BY Regiao
                ORDER BY faturamento_total DESC
            """
            return self.db.execute_query(query) or []

        except Exception as e:
            logging.error(f"❌ Erro ao gerar relatório região: {e}")
            return []

    def gerar_relatorio_inventario(self):
        """Gera relatório completo do inventário"""
        try:
            query = """
                SELECT Codigo_Ativo, Endereco, Regiao, Tipo_Placa, Status_Atual,
                       Cliente_Locacao, Valor_Mensal, Data_Cadastro
                FROM placas
                ORDER BY Regiao, Codigo_Ativo
            """
            return self.db.execute_query(query) or []

        except Exception as e:
            logging.error(f"❌ Erro ao gerar relatório inventário: {e}")
            return []

# =============================================================================
# INICIALIZAR MANAGERS
# =============================================================================
relatorio_manager = RelatorioManager(DatabaseConnection())

# =============================================================================
# ROTAS PARA MANIPULAÇÃO DE PDFs
# =============================================================================
@app.route('/uploads/contratos/<filename>')
def servir_pdf(filename):
    """Serve arquivos PDF da pasta de uploads"""
    try:
        return send_from_directory(app.config['PDF_FOLDER'], filename)
    except FileNotFoundError:
        return "Arquivo não encontrado", 404

@app.route('/contratos/pdf/<int:contrato_id>')
def visualizar_pdf(contrato_id):
    """Visualiza PDF de um contrato específico"""
    if 'user' not in session:
        return redirect('/login')

    try:
        db = DatabaseConnection()
        contrato = db.execute_query("SELECT arquivo_pdf FROM contratos WHERE id = %s", (contrato_id,))

        if not contrato or not contrato[0]['arquivo_pdf']:
            return "Contrato ou arquivo PDF não encontrado", 404

        filename = contrato[0]['arquivo_pdf']
        filepath = os.path.join(app.config['PDF_FOLDER'], filename)

        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=False)
        else:
            # Se o arquivo não existe fisicamente, criar um PDF simulado
            return criar_pdf_simulado(filename, f"Contrato ID: {contrato_id}")

    except Exception as e:
        logging.error(f"❌ Erro ao visualizar PDF: {e}")
        return "Erro ao carregar PDF", 500

def criar_pdf_simulado(filename, conteudo):
    """Cria um PDF simulado para demonstração"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.drawString(100, 750, f"CONTRATO - {conteudo}")
    p.drawString(100, 730, "Este é um PDF simulado para demonstração.")
    p.drawString(100, 710, f"Arquivo: {filename}")
    p.drawString(100, 690, "Em ambiente real, este seria o contrato real.")
    p.drawString(100, 670, f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    p.showPage()
    p.save()

    buffer.seek(0)
    return send_file(buffer, as_attachment=False, download_name=filename, mimetype='application/pdf')

# =============================================================================
# ROTAS PRINCIPAIS
# =============================================================================
@app.route('/')
def index():
    """Página inicial - redireciona para login"""
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']

        # Simulação de login bem-sucedido
        if email in ['admin@formato.com', 'eduardo@formato.com', 'pamela@formato.com'] and senha == 'teste123':
            session['user'] = {
                'nome': 'Usuário Teste',
                'perfil': 'Administrador'
            }
            session.permanent = True
            return redirect('/inventario')
        else:
            return '''
            <!DOCTYPE html>
            <html>
            <head>
                <title>Login - SIG-ME</title>
                <style>
                    body { font-family: Arial, sans-serif; padding: 40px; text-align: center; }
                    .login-box { max-width: 400px; margin: 0 auto; padding: 40px; border: 1px solid #ddd; border-radius: 10px; }
                    input, button { width: 100%; padding: 10px; margin: 10px 0; }
                    .error { color: red; margin-top: 10px; }
                </style>
            </head>
            <body>
                <div class="login-box">
                    <h1>🚀 SIG-ME</h1>
                    <h3>Login</h3>
                    <div class="error">Login falhou! Verifique suas credenciais.</div>
                    <form method="POST">
                        <input type="email" name="email" placeholder="Email" required>
                        <input type="password" name="senha" placeholder="Senha" required>
                        <button type="submit">Entrar</button>
                    </form>
                    <p><strong>Contas teste:</strong> admin@formato.com / teste123</p>
                </div>
            </body>
            </html>
            '''

    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Login - SIG-ME</title>
        <style>
            body { font-family: Arial, sans-serif; padding: 40px; text-align: center; }
            .login-box { max-width: 400px; margin: 0 auto; padding: 40px; border: 1px solid #ddd; border-radius: 10px; }
            input, button { width: 100%; padding: 10px; margin: 10px 0; }
        </style>
    </head>
    <body>
        <div class="login-box">
            <h1>🚀 SIG-ME</h1>
            <h3>Login</h3>
            <form method="POST">
                <input type="email" name="email" placeholder="Email" required>
                <input type="password" name="senha" placeholder="Senha" required>
                <button type="submit">Entrar</button>
            </form>
            <p><strong>Contas teste:</strong> admin@formato.com / teste123</p>
        </div>
    </body>
    </html>
    '''

@app.route('/logout')
def logout():
    """Logout do usuário"""
    session.clear()
    return redirect('/login')

# =============================================================================
# MÓDULO DE INVENTÁRIO (COMPLETO)
# =============================================================================
@app.route('/inventario')
def inventario():
    """Página principal do inventário"""
    if 'user' not in session:
        return redirect('/login')

    db = DatabaseConnection()

    # Buscar todas as placas
    placas = db.execute_query("""
        SELECT * FROM placas
        ORDER BY Regiao, Codigo_Ativo
    """) or []

    # Métricas do inventário
    metricas = db.execute_query("""
        SELECT
            COUNT(*) as total_placas,
            SUM(CASE WHEN Status_Atual = 'disponível' THEN 1 ELSE 0 END) as disponiveis,
            SUM(CASE WHEN Status_Atual = 'locado' THEN 1 ELSE 0 END) as locadas,
            SUM(CASE WHEN Status_Atual = 'reservado' THEN 1 ELSE 0 END) as reservadas,
            SUM(CASE WHEN Status_Atual = 'manutenção' THEN 1 ELSE 0 END) as manutencao,
            SUM(Valor_Mensal) as faturamento_potencial
        FROM placas
    """)
    metricas = metricas[0] if metricas else {}

    # Distribuição por região
    distribuicao_regiao = db.execute_query("""
        SELECT Regiao, COUNT(*) as quantidade
        FROM placas
        GROUP BY Regiao
        ORDER BY quantidade DESC
    """) or []

    return render_inventario_template(session['user'], placas, metricas, distribuicao_regiao)

@app.route('/inventario/placa/nova', methods=['POST'])
def nova_placa():
    """Cria uma nova placa - CORRIGIDO"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        dados = request.get_json()
        db = DatabaseConnection()

        # CORREÇÃO: Query atualizada para incluir todos os campos
        query = """
            INSERT INTO placas
            (Codigo_Ativo, Endereco, Regiao, Tipo_Placa, Valor_Mensal, Data_Cadastro, Status_Atual, Cliente_Locacao)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        params = (
            dados['codigo_ativo'],
            dados['endereco'],
            dados['regiao'],
            dados['tipo_placa'],
            dados['valor_mensal'],
            datetime.datetime.now().date(),
            dados['status_atual'],  # CORREÇÃO: Usar status do formulário
            dados.get('cliente_locacao')  # CORREÇÃO: Incluir cliente
        )

        success = db.execute_update(query, params)

        if success:
            return jsonify({'success': True, 'message': 'Placa criada com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao criar placa'})

    except Exception as e:
        logging.error(f"❌ Erro ao criar placa: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/inventario/placa/editar', methods=['POST'])
def editar_placa():
    """Edita uma placa existente - CORRIGIDO"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        dados = request.get_json()
        db = DatabaseConnection()

        # CORREÇÃO: Query atualizada para garantir que todos os campos sejam atualizados
        query = """
            UPDATE placas
            SET Endereco = %s, Regiao = %s, Tipo_Placa = %s, Valor_Mensal = %s,
                Status_Atual = %s, Cliente_Locacao = %s
            WHERE Codigo_Ativo = %s
        """
        params = (
            dados['endereco'],
            dados['regiao'],
            dados['tipo_placa'],
            dados['valor_mensal'],
            dados['status_atual'],
            dados.get('cliente_locacao'),  # CORREÇÃO: Usar get() para evitar KeyError
            dados['codigo_ativo']
        )

        success = db.execute_update(query, params)

        if success:
            return jsonify({'success': True, 'message': 'Placa atualizada com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao atualizar placa'})

    except Exception as e:
        logging.error(f"❌ Erro ao editar placa: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/inventario/placa/excluir/<string:codigo_placa>', methods=['DELETE'])
def excluir_placa(codigo_placa):
    """Exclui uma placa"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        db = DatabaseConnection()

        query = "DELETE FROM placas WHERE Codigo_Ativo = %s"
        success = db.execute_update(query, (codigo_placa,))

        if success:
            return jsonify({'success': True, 'message': 'Placa excluída com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao excluir placa'})

    except Exception as e:
        logging.error(f"❌ Erro ao excluir placa: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/inventario/placa/<string:codigo_placa>')
def detalhes_placa(codigo_placa):
    """Detalhes de uma placa específica"""
    if 'user' not in session:
        return redirect('/login')

    db = DatabaseConnection()
    placa = db.execute_query("SELECT * FROM placas WHERE Codigo_Ativo = %s", (codigo_placa,))

    if not placa:
        return "Placa não encontrada", 404

    placa = placa[0]

    # HTML para detalhes da placa
    return f'''
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Detalhes da Placa - SIG-ME</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: {CORES_FORMATO['claro']};
            color: {CORES_FORMATO['texto']};
        }}
        .header {{
            background: {CORES_FORMATO['branco']};
            padding: 20px 30px;
            box-shadow: 0 2px 20px rgba(0, 0, 0, 0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 4px solid {CORES_FORMATO['destaque']};
        }}
        .logo h1 {{
            color: {CORES_FORMATO['primaria']};
            font-size: 24px;
            font-weight: 700;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
            color: {CORES_FORMATO['texto']};
            font-weight: 500;
        }}
        .btn-logout {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
        }}
        .btn-logout:hover {{
            background: {CORES_FORMATO['alerta']};
            transform: translateY(-2px);
        }}
        .container {{
            max-width: 1200px;
            margin: 30px auto;
            padding: 0 25px;
        }}
        .navigation {{
            display: flex;
            gap: 12px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .nav-btn {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 14px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .nav-btn:hover {{
            background: {CORES_FORMATO['secundaria']};
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(26, 54, 93, 0.2);
        }}
        .content {{
            background: {CORES_FORMATO['branco']};
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        }}
        .placa-detalhes {{
            max-width: 600px;
            margin: 0 auto;
        }}
        .placa-header {{
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .placa-codigo {{
            font-size: 32px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 10px;
        }}
        .placa-status {{
            display: inline-block;
            padding: 8px 16px;
            border-radius: 20px;
            font-weight: 600;
            font-size: 14px;
        }}
        .status-disponivel {{ background: {CORES_FORMATO['sucesso']}20; color: {CORES_FORMATO['sucesso']}; }}
        .status-locado {{ background: {CORES_FORMATO['destaque']}20; color: {CORES_FORMATO['destaque']}; }}
        .status-reservado {{ background: {CORES_FORMATO['alerta']}20; color: {CORES_FORMATO['alerta']}; }}
        .status-manutencao {{ background: {CORES_FORMATO['info']}20; color: {CORES_FORMATO['info']}; }}
        .placa-info {{
            margin-bottom: 30px;
        }}
        .info-item {{
            margin-bottom: 15px;
            padding: 15px;
            background: {CORES_FORMATO['claro']};
            border-radius: 8px;
        }}
        .info-label {{
            font-weight: 600;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 5px;
        }}
        .info-value {{
            color: {CORES_FORMATO['texto']};
        }}
        .btn-voltar {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            display: inline-block;
            font-weight: 600;
        }}
        .btn-salvar {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            margin-left: 10px;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-group label {{
            display: block;
            margin-bottom: 5px;
            font-weight: 600;
            color: {CORES_FORMATO['primaria']};
        }}
        .form-group input, .form-group select {{
            width: 100%;
            padding: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
            border-radius: 5px;
            font-size: 14px;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <h1>🚀 SIG-ME - Detalhes da Placa</h1>
        </div>
        <div class="user-info">
            <span>👋 Olá, {session['user']['nome']} <small>({session['user']['perfil']})</small></span>
            <a href="/logout" class="btn-logout">🚪 Sair</a>
        </div>
    </div>

    <div class="container">
        <div class="navigation">
            <a href="/inventario" class="nav-btn">📋 Voltar ao Inventário</a>
        </div>

        <div class="content">
            <div class="placa-detalhes">
                <div class="placa-header">
                    <div class="placa-codigo">{placa['Codigo_Ativo']}</div>
                    <div class="placa-status status-{placa['Status_Atual'].lower().replace('ç', 'c').replace('ã', 'a')}">{placa['Status_Atual']}</div>
                </div>

                <form id="formEditarPlaca">
                    <div class="placa-info">
                        <div class="form-group">
                            <label for="endereco">📍 Endereço</label>
                            <input type="text" id="endereco" name="endereco" value="{placa['Endereco']}" required>
                        </div>
                        <div class="form-group">
                            <label for="regiao">🏙️ Região</label>
                            <select id="regiao" name="regiao" required>
                                <option value="São Roque" {'selected' if placa['Regiao'] == 'São Roque' else ''}>São Roque</option>
                                <option value="Mairinque" {'selected' if placa['Regiao'] == 'Mairinque' else ''}>Mairinque</option>
                                <option value="Ibuna" {'selected' if placa['Regiao'] == 'Ibuna' else ''}>Ibuna</option>
                                <option value="Araçá" {'selected' if placa['Regiao'] == 'Araçá' else ''}>Araçá</option>
                                <option value="Piedade" {'selected' if placa['Regiao'] == 'Piedade' else ''}>Piedade</option>
                                <option value="Alumínio" {'selected' if placa['Regiao'] == 'Alumínio' else ''}>Alumínio</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label for="tipo_placa">📺 Tipo de Placa</label>
                            <select id="tipo_placa" name="tipo_placa" required>
                                <option value="Outdoor" {'selected' if placa['Tipo_Placa'] == 'Outdoor' else ''}>Outdoor</option>
                                <option value="Front Light" {'selected' if placa['Tipo_Placa'] == 'Front Light' else ''}>Front Light</option>
                                <option value="Back Light" {'selected' if placa['Tipo_Placa'] == 'Back Light' else ''}>Back Light</option>
                                <option value="Led" {'selected' if placa['Tipo_Placa'] == 'Led' else ''}>Led</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label for="cliente_locacao">👥 Cliente</label>
                            <input type="text" id="cliente_locacao" name="cliente_locacao" value="{placa['Cliente_Locacao'] or ''}" placeholder="Deixe vazio se não houver cliente">
                        </div>
                        <div class="form-group">
                            <label for="valor_mensal">💰 Valor Mensal (R$)</label>
                            <input type="number" id="valor_mensal" name="valor_mensal" value="{placa['Valor_Mensal']}" step="0.01" min="0" required>
                        </div>
                        <div class="form-group">
                            <label for="status_atual">Status</label>
                            <select id="status_atual" name="status_atual" required>
                                <option value="disponível" {'selected' if placa['Status_Atual'] == 'disponível' else ''}>Disponível</option>
                                <option value="locado" {'selected' if placa['Status_Atual'] == 'locado' else ''}>Locado</option>
                                <option value="reservado" {'selected' if placa['Status_Atual'] == 'reservado' else ''}>Reservado</option>
                                <option value="manutenção" {'selected' if placa['Status_Atual'] == 'manutenção' else ''}>Manutenção</option>
                            </select>
                        </div>
                    </div>

                    <div style="text-align: center; margin-top: 30px;">
                        <a href="/inventario" class="btn-voltar">← Voltar ao Inventário</a>
                        <button type="submit" class="btn-salvar">💾 Salvar Alterações</button>
                    </div>
                </form>
            </div>
        </div>
    </div>

    <script>
        document.getElementById('formEditarPlaca').addEventListener('submit', function(e) {{
            e.preventDefault();

            const formData = new FormData(this);
            const dados = {{
                codigo_ativo: '{placa['Codigo_Ativo']}',
                endereco: formData.get('endereco'),
                regiao: formData.get('regiao'),
                tipo_placa: formData.get('tipo_placa'),
                valor_mensal: parseFloat(formData.get('valor_mensal')),
                status_atual: formData.get('status_atual'),
                cliente_locacao: formData.get('cliente_locacao') || null
            }};

            fetch('/inventario/placa/editar', {{
                method: 'POST',
                headers: {{
                    'Content-Type': 'application/json',
                }},
                body: JSON.stringify(dados)
            }})
            .then(response => response.json())
            .then(data => {{
                if (data.success) {{
                    alert('Placa atualizada com sucesso!');
                    window.location.href = '/inventario';
                }} else {{
                    alert('Erro ao atualizar placa: ' + data.message);
                }}
            }})
            .catch(error => {{
                console.error('Erro:', error);
                alert('Erro ao atualizar placa');
            }});
        }});
    </script>
</body>
</html>
'''

def render_inventario_template(user, placas, metricas, distribuicao_regiao):
    """Renderiza o template de inventário completo"""

    metricas_html = f"""
    <div class="metrics-grid">
        <div class="metric-card">
            <div class="metric-value">{metricas.get('total_placas', 0)}</div>
            <div class="metric-label">Total de Placas</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{metricas.get('disponiveis', 0)}</div>
            <div class="metric-label">Disponíveis</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{metricas.get('locadas', 0)}</div>
            <div class="metric-label">Locadas</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{metricas.get('reservadas', 0)}</div>
            <div class="metric-label">Reservadas</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{metricas.get('manutencao', 0)}</div>
            <div class="metric-label">Manutenção</div>
        </div>
    </div>
    """

    distribuicao_html = ""
    for regiao in distribuicao_regiao:
        distribuicao_html += f"""
        <div class="distribuicao-item" onclick="filtrarPorRegiao('{regiao['Regiao']}')">
            <div class="regiao-nome">{regiao['Regiao']}</div>
            <div class="regiao-quantidade">{regiao['quantidade']} placas</div>
        </div>
        """

    placas_html = ""
    for placa in placas:
        status_class = placa['Status_Atual'].lower().replace('ç', 'c').replace('ã', 'a')
        cliente = placa['Cliente_Locacao'] or 'Sem locação'
        placas_html += f"""
        <div class="placa-card {status_class}">
            <div class="placa-header">
                <h4>📋 {placa['Codigo_Ativo']}</h4>
                <span class="status-badge {status_class}">{placa['Status_Atual']}</span>
            </div>
            <div class="placa-info">
                <p><strong>📍 Endereço:</strong> {placa['Endereco']}</p>
                <p><strong>🏙️ Região:</strong> {placa['Regiao']}</p>
                <p><strong>📺 Tipo:</strong> {placa['Tipo_Placa']}</p>
                <p><strong>👥 Cliente:</strong> {cliente}</p>
                <p><strong>💰 Valor:</strong> R$ {placa['Valor_Mensal']:,.2f}</p>
                <p><strong>📅 Cadastro:</strong> {placa['Data_Cadastro']}</p>
            </div>
            <div class="placa-actions">
                <button class="btn-action" onclick="editarPlaca('{placa['Codigo_Ativo']}')">✏️ Editar</button>
                <button class="btn-action" onclick="detalhesPlaca('{placa['Codigo_Ativo']}')">👁️ Detalhes</button>
                <button class="btn-action btn-excluir" onclick="excluirPlaca('{placa['Codigo_Ativo']}')">🗑️ Excluir</button>
            </div>
        </div>
        """

    def convert_to_serializable(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, datetime.date):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {key: convert_to_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [convert_to_serializable(item) for item in obj]
        else:
            return obj

    placas_serializable = convert_to_serializable(placas)

    # CORREÇÃO: Template string JavaScript corrigida
    template_js = """
    placasFiltradas.forEach(placa => {
        const status_class = placa.Status_Atual.toLowerCase().replace('ç', 'c').replace('ã', 'a');
        const cliente = placa.Cliente_Locacao || 'Sem locação';

        placasGrid.innerHTML += `
        <div class="placa-card ${status_class}">
            <div class="placa-header">
                <h4>📋 ${placa.Codigo_Ativo}</h4>
                <span class="status-badge ${status_class}">${placa.Status_Atual}</span>
            </div>
            <div class="placa-info">
                <p><strong>📍 Endereço:</strong> ${placa.Endereco}</p>
                <p><strong>🏙️ Região:</strong> ${placa.Regiao}</p>
                <p><strong>📺 Tipo:</strong> ${placa.Tipo_Placa}</p>
                <p><strong>👥 Cliente:</strong> ${cliente}</p>
                <p><strong>💰 Valor:</strong> R$ ${placa.Valor_Mensal.toFixed(2).replace('.', ',')}</p>
                <p><strong>📅 Cadastro:</strong> ${placa.Data_Cadastro}</p>
            </div>
            <div class="placa-actions">
                <button class="btn-action" onclick="editarPlaca('${placa.Codigo_Ativo}')">✏️ Editar</button>
                <button class="btn-action" onclick="detalhesPlaca('${placa.Codigo_Ativo}')">👁️ Detalhes</button>
                <button class="btn-action btn-excluir" onclick="excluirPlaca('${placa.Codigo_Ativo}')">🗑️ Excluir</button>
            </div>
        </div>
        `;
    });
    """

    return f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Inventário - SIG-ME</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: {CORES_FORMATO['claro']};
            color: {CORES_FORMATO['texto']};
        }}
        .header {{
            background: {CORES_FORMATO['branco']};
            padding: 20px 30px;
            box-shadow: 0 2px 20px rgba(0, 0, 0, 0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 4px solid {CORES_FORMATO['destaque']};
        }}
        .logo h1 {{
            color: {CORES_FORMATO['primaria']};
            font-size: 24px;
            font-weight: 700;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
            color: {CORES_FORMATO['texto']};
            font-weight: 500;
        }}
        .btn-logout {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
        }}
        .btn-logout:hover {{
            background: {CORES_FORMATO['alerta']};
            transform: translateY(-2px);
        }}
        .container {{
            max-width: 1400px;
            margin: 30px auto;
            padding: 0 25px;
        }}
        .navigation {{
            display: flex;
            gap: 12px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .nav-btn {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 14px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .nav-btn:hover {{
            background: {CORES_FORMATO['secundaria']};
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(26, 54, 93, 0.2);
        }}
        .nav-btn.active {{
            background: {CORES_FORMATO['destaque']};
        }}
        .content {{
            background: {CORES_FORMATO['branco']};
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        }}
        .section-title {{
            font-size: 24px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        .metric-card {{
            background: {CORES_FORMATO['claro']};
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            border-left: 4px solid {CORES_FORMATO['info']};
        }}
        .metric-value {{
            font-size: 28px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 5px;
        }}
        .metric-label {{
            color: {CORES_FORMATO['texto_claro']};
            font-size: 14px;
            font-weight: 600;
        }}
        .grid-2col {{
            display: grid;
            grid-template-columns: 1fr 2fr;
            gap: 30px;
            margin-top: 30px;
        }}
        .card {{
            background: {CORES_FORMATO['claro']};
            padding: 25px;
            border-radius: 10px;
        }}
        .card h3 {{
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            font-size: 18px;
        }}
        .distribuicao-item {{
            background: {CORES_FORMATO['branco']};
            padding: 15px;
            border-radius: 6px;
            margin-bottom: 10px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
            transition: all 0.3s;
        }}
        .distribuicao-item:hover {{
            background: {CORES_FORMATO['info']}20;
            transform: translateX(5px);
        }}
        .regiao-nome {{
            font-weight: 600;
            color: {CORES_FORMATO['primaria']};
        }}
        .regiao-quantidade {{
            color: {CORES_FORMATO['texto_claro']};
            font-size: 14px;
        }}
        .placas-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(350px, 1fr));
            gap: 20px;
            max-height: 600px;
            overflow-y: auto;
            padding: 10px;
        }}
        .placa-card {{
            background: {CORES_FORMATO['branco']};
            padding: 20px;
            border-radius: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
        }}
        .placa-card.disponivel {{ border-left: 4px solid {CORES_FORMATO['sucesso']}; }}
        .placa-card.locado {{ border-left: 4px solid {CORES_FORMATO['destaque']}; }}
        .placa-card.reservado {{ border-left: 4px solid {CORES_FORMATO['alerta']}; }}
        .placa-card.manutencao {{ border-left: 4px solid {CORES_FORMATO['info']}; }}
        .placa-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 1px solid {CORES_FORMATO['claro']};
        }}
        .placa-header h4 {{
            color: {CORES_FORMATO['primaria']};
        }}
        .status-badge {{
            padding: 6px 12px;
            border-radius: 15px;
            font-size: 12px;
            font-weight: 600;
        }}
        .status-badge.disponivel {{ background: {CORES_FORMATO['sucesso']}20; color: {CORES_FORMATO['sucesso']}; }}
        .status-badge.locado {{ background: {CORES_FORMATO['destaque']}20; color: {CORES_FORMATO['destaque']}; }}
        .status-badge.reservado {{ background: {CORES_FORMATO['alerta']}20; color: {CORES_FORMATO['alerta']}; }}
        .status-badge.manutencao {{ background: {CORES_FORMATO['info']}20; color: {CORES_FORMATO['info']}; }}
        .placa-info p {{
            margin: 8px 0;
            font-size: 14px;
        }}
        .placa-actions {{
            display: flex;
            gap: 10px;
            margin-top: 15px;
        }}
        .btn-action {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            border: none;
            padding: 8px 15px;
            border-radius: 5px;
            cursor: pointer;
            font-size: 12px;
            flex: 1;
        }}
        .btn-action:hover {{
            background: {CORES_FORMATO['secundaria']};
        }}
        .btn-excluir {{
            background: {CORES_FORMATO['destaque']};
        }}
        .btn-excluir:hover {{
            background: #c53030;
        }}
        .btn-nova-placa {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            margin-bottom: 20px;
            transition: all 0.3s;
        }}
        .btn-nova-placa:hover {{
            background: {CORES_FORMATO['info']};
            transform: translateY(-2px);
        }}
        .alert {{
            padding: 15px;
            margin: 15px 0;
            border-radius: 5px;
            font-weight: 500;
        }}
        .alert.success {{
            background: {CORES_FORMATO['sucesso']}20;
            color: {CORES_FORMATO['sucesso']};
            border: 1px solid {CORES_FORMATO['sucesso']};
        }}
        .alert.error {{
            background: {CORES_FORMATO['destaque']}20;
            color: {CORES_FORMATO['destaque']};
            border: 1px solid {CORES_FORMATO['destaque']};
        }}

        /* MODAL STYLES - ATUALIZADO COM SCROLL E BOTÃO SALVAR */
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow-y: auto;
        }}
        .modal-content {{
            background-color: {CORES_FORMATO['branco']};
            margin: 5% auto;
            padding: 30px;
            border-radius: 10px;
            width: 90%;
            max-width: 500px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.2);
            max-height: 85vh;
            overflow-y: auto;
        }}
        .modal-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .close {{
            color: {CORES_FORMATO['texto_claro']};
            font-size: 28px;
            font-weight: bold;
            cursor: pointer;
        }}
        .close:hover {{
            color: {CORES_FORMATO['destaque']};
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-group label {{
            display: block;
            margin-bottom: 5px;
            font-weight: 600;
            color: {CORES_FORMATO['primaria']};
        }}
        .form-group input, .form-group select {{
            width: 100%;
            padding: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
            border-radius: 5px;
            font-size: 14px;
        }}
        .form-actions {{
            display: flex;
            gap: 10px;
            justify-content: flex-end;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid {CORES_FORMATO['claro']};
        }}
        .btn-cancelar {{
            background: {CORES_FORMATO['texto_claro']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: 600;
        }}
        .btn-salvar {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: 600;
        }}
        .btn-cancelar:hover {{
            background: {CORES_FORMATO['secundaria']};
        }}
        .btn-salvar:hover {{
            background: #2f855a;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <h1>🚀 SIG-ME - Inventário</h1>
        </div>
        <div class="user-info">
            <span>👋 Olá, {user['nome']} <small>({user['perfil']})</small></span>
            <a href="/logout" class="btn-logout">🚪 Sair</a>
        </div>
    </div>

    <div class="container">
        <div class="navigation">
            <a href="/inventario" class="nav-btn active">📋 Inventário</a>
            <a href="/relatorios" class="nav-btn">📊 Relatórios</a>
            <a href="/comercial" class="nav-btn">💼 Comercial</a>
            <a href="/contratos" class="nav-btn">📑 Contratos</a>
            <a href="/exportar_dados" class="nav-btn">📥 Exportar Dados</a>
        </div>

        <div class="content">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                <h2 class="section-title">📋 Módulo de Inventário</h2>
                <button class="btn-nova-placa" onclick="abrirModalNovaPlaca()">➕ Nova Placa</button>
            </div>

            <div id="alert-container"></div>

            {metricas_html}

            <div class="grid-2col">
                <div class="card">
                    <h3>🏙️ Distribuição por Região</h3>
                    {distribuicao_html if distribuicao_html else '<p>Nenhuma placa cadastrada</p>'}
                </div>

                <div class="card">
                    <h3>📋 Lista de Placas</h3>
                    <div class="placas-grid" id="placas-grid">
                        {placas_html if placas_html else '<p>Nenhuma placa cadastrada</p>'}
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Modal Nova Placa - ATUALIZADO COM SCROLL E BOTÃO SALVAR -->
    <div id="modalPlaca" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h3 id="modalTituloPlaca">Nova Placa</h3>
                <span class="close" onclick="fecharModalPlaca()">&times;</span>
            </div>
            <form id="formPlaca">
                <input type="hidden" id="placa_codigo_editar" name="placa_codigo_editar">

                <div class="form-group">
                    <label for="codigo_ativo"><strong>Código da Placa:</strong></label>
                    <input type="text" id="codigo_ativo" name="codigo_ativo" required>
                </div>

                <div class="form-group">
                    <label for="endereco"><strong>Endereço:</strong></label>
                    <input type="text" id="endereco" name="endereco" required>
                </div>

                <div class="form-group">
                    <label for="regiao"><strong>Região:</strong></label>
                    <select id="regiao" name="regiao" required>
                        <option value="">Selecione a região</option>
                        <option value="São Roque">São Roque</option>
                        <option value="Mairinque">Mairinque</option>
                        <option value="Ibuna">Ibuna</option>
                        <option value="Araçá">Araçá</option>
                        <option value="Piedade">Piedade</option>
                        <option value="Alumínio">Alumínio</option>
                    </select>
                </div>

                <div class="form-group">
                    <label for="tipo_placa"><strong>Tipo de Placa:</strong></label>
                    <select id="tipo_placa" name="tipo_placa" required>
                        <option value="">Selecione o tipo</option>
                        <option value="Outdoor">Outdoor</option>
                        <option value="Front Light">Front Light</option>
                        <option value="Back Light">Back Light</option>
                        <option value="Led">Led</option>
                    </select>
                </div>

                <div class="form-group">
                    <label for="valor_mensal"><strong>Valor Mensal (R$):</strong></label>
                    <input type="number" id="valor_mensal" name="valor_mensal" step="0.01" min="0" required>
                </div>

                <div class="form-group">
                    <label for="status_atual"><strong>Status:</strong></label>
                    <select id="status_atual" name="status_atual" required>
                        <option value="disponível">Disponível</option>
                        <option value="locado">Locado</option>
                        <option value="reservado">Reservado</option>
                        <option value="manutenção">Manutenção</option>
                    </select>
                </div>

                <div class="form-group">
                    <label for="cliente_locacao"><strong>Cliente:</strong></label>
                    <input type="text" id="cliente_locacao" name="cliente_locacao" placeholder="Deixe vazio se não houver cliente">
                </div>

                <div class="form-actions">
                    <button type="button" class="btn-cancelar" onclick="fecharModalPlaca()">Cancelar</button>
                    <button type="submit" class="btn-salvar">💾 Salvar Placa</button>
                </div>
            </form>
        </div>
    </div>

    <script>
        let placaEditando = null;
        const placas = {json.dumps(placas_serializable)};

        function abrirModalNovaPlaca() {{
            document.getElementById('modalTituloPlaca').textContent = 'Nova Placa';
            document.getElementById('formPlaca').reset();
            document.getElementById('placa_codigo_editar').value = '';
            document.getElementById('codigo_ativo').disabled = false;
            placaEditando = null;
            document.getElementById('modalPlaca').style.display = 'block';
        }}

        function fecharModalPlaca() {{
            document.getElementById('modalPlaca').style.display = 'none';
        }}

        function editarPlaca(codigo) {{
            const placa = placas.find(p => p.Codigo_Ativo === codigo);

            if (placa) {{
                document.getElementById('modalTituloPlaca').textContent = 'Editar Placa';
                document.getElementById('placa_codigo_editar').value = placa.Codigo_Ativo;
                document.getElementById('codigo_ativo').value = placa.Codigo_Ativo;
                document.getElementById('codigo_ativo').disabled = true;
                document.getElementById('endereco').value = placa.Endereco;
                document.getElementById('regiao').value = placa.Regiao;
                document.getElementById('tipo_placa').value = placa.Tipo_Placa;
                document.getElementById('valor_mensal').value = placa.Valor_Mensal;
                document.getElementById('status_atual').value = placa.Status_Atual;
                document.getElementById('cliente_locacao').value = placa.Cliente_Locacao || '';

                placaEditando = placa;
                document.getElementById('modalPlaca').style.display = 'block';
            }}
        }}

        function detalhesPlaca(codigo) {{
            window.location.href = '/inventario/placa/' + codigo;
        }}

        function excluirPlaca(codigo) {{
            if (confirm('Tem certeza que deseja excluir a placa ' + codigo + '?')) {{
                fetch('/inventario/placa/excluir/' + codigo, {{
                    method: 'DELETE'
                }})
                .then(response => response.json())
                .then(data => {{
                    if (data.success) {{
                        mostrarAlert(data.message, 'success');
                        setTimeout(() => location.reload(), 1500);
                    }} else {{
                        mostrarAlert(data.message, 'error');
                    }}
                }})
                .catch(error => {{
                    console.error('Erro:', error);
                    mostrarAlert('Erro ao excluir placa', 'error');
                }});
            }}
        }}

        function filtrarPorRegiao(regiao) {{
            const placasGrid = document.getElementById('placas-grid');
            placasGrid.innerHTML = '';

            const placasFiltradas = placas.filter(placa => placa.Regiao === regiao);

            if (placasFiltradas.length === 0) {{
                placasGrid.innerHTML = '<p>Nenhuma placa encontrada para esta região.</p>';
                return;
            }}

            placasFiltradas.forEach(placa => {{
                const status_class = placa.Status_Atual.toLowerCase().replace('ç', 'c').replace('ã', 'a');
                const cliente = placa.Cliente_Locacao || 'Sem locação';

                placasGrid.innerHTML += `
                <div class="placa-card ${{status_class}}">
                    <div class="placa-header">
                        <h4>📋 ${{placa.Codigo_Ativo}}</h4>
                        <span class="status-badge ${{status_class}}">${{placa.Status_Atual}}</span>
                    </div>
                    <div class="placa-info">
                        <p><strong>📍 Endereço:</strong> ${{placa.Endereco}}</p>
                        <p><strong>🏙️ Região:</strong> ${{placa.Regiao}}</p>
                        <p><strong>📺 Tipo:</strong> ${{placa.Tipo_Placa}}</p>
                        <p><strong>👥 Cliente:</strong> ${{cliente}}</p>
                        <p><strong>💰 Valor:</strong> R$ ${{placa.Valor_Mensal.toFixed(2).replace('.', ',')}}</p>
                        <p><strong>📅 Cadastro:</strong> ${{placa.Data_Cadastro}}</p>
                    </div>
                    <div class="placa-actions">
                        <button class="btn-action" onclick="editarPlaca('${{placa.Codigo_Ativo}}')">✏️ Editar</button>
                        <button class="btn-action" onclick="detalhesPlaca('${{placa.Codigo_Ativo}}')">👁️ Detalhes</button>
                        <button class="btn-action btn-excluir" onclick="excluirPlaca('${{placa.Codigo_Ativo}}')">🗑️ Excluir</button>
                    </div>
                </div>
                `;
            }});
        }}

        function mostrarAlert(mensagem, tipo) {{
            const alertContainer = document.getElementById('alert-container');
            const alert = document.createElement('div');
            alert.className = 'alert ' + tipo;
            alert.textContent = mensagem;
            alertContainer.appendChild(alert);

            setTimeout(() => {{
                alert.remove();
            }}, 5000);
        }}

        // Fechar modal ao clicar fora
        window.onclick = function(event) {{
            const modal = document.getElementById('modalPlaca');
            if (event.target === modal) {{
                fecharModalPlaca();
            }}
        }}

        // Envio do formulário - CORREÇÃO: Garantir que todos os dados sejam enviados
        document.getElementById('formPlaca').addEventListener('submit', function(e) {{
            e.preventDefault();

            const formData = new FormData(this);
            const codigoEditar = formData.get('placa_codigo_editar');

            const dados = {{
                codigo_ativo: formData.get('codigo_ativo'),
                endereco: formData.get('endereco'),
                regiao: formData.get('regiao'),
                tipo_placa: formData.get('tipo_placa'),
                valor_mensal: parseFloat(formData.get('valor_mensal')),
                status_atual: formData.get('status_atual'),
                cliente_locacao: formData.get('cliente_locacao') || null  // CORREÇÃO: Garantir null se vazio
            }};

            const url = codigoEditar ? '/inventario/placa/editar' : '/inventario/placa/nova';
            const method = 'POST';

            fetch(url, {{
                method: method,
                headers: {{
                    'Content-Type': 'application/json',
                }},
                body: JSON.stringify(dados)
            }})
            .then(response => response.json())
            .then(data => {{
                if (data.success) {{
                    mostrarAlert(data.message, 'success');
                    fecharModalPlaca();
                    setTimeout(() => location.reload(), 1500);
                }} else {{
                    mostrarAlert(data.message, 'error');
                }}
            }})
            .catch(error => {{
                console.error('Erro:', error);
                mostrarAlert('Erro ao salvar placa', 'error');
            }});
        }});
    </script>
</body>
</html>
'''

# =============================================================================
# MÓDULO COMERCIAL (CORRIGIDO)
# =============================================================================
@app.route('/comercial')
def comercial():
    """Página comercial"""
    if 'user' not in session:
        return redirect('/login')

    db = DatabaseConnection()

    # CORREÇÃO: Query simplificada para métricas
    metricas_comerciais = db.execute_query("""
        SELECT
            COUNT(*) as total_placas,
            SUM(CASE WHEN Status_Atual IN ('locado', 'reservado') THEN Valor_Mensal ELSE 0 END) as faturamento_mensal,
            SUM(CASE WHEN Status_Atual IN ('locado', 'reservado') THEN 1 ELSE 0 END) as placas_ocupadas,
            SUM(CASE WHEN Status_Atual = 'disponível' THEN 1 ELSE 0 END) as placas_disponiveis,
            COUNT(DISTINCT Cliente_Locacao) as total_clientes
        FROM placas
    """)
    metricas_comerciais = metricas_comerciais[0] if metricas_comerciais else {}

    # CORREÇÃO: Query corrigida para clientes - incluindo ID explicitamente
    clientes = db.execute_query("SELECT id, Nome_Fantasia, Nome_Razao_Social, CNPJ_CPF, Telefone, Email FROM clientes ORDER BY Nome_Fantasia") or []

    return render_comercial_template(session['user'], metricas_comerciais, clientes, [])

@app.route('/comercial/cliente/novo', methods=['POST'])
def novo_cliente():
    """Cria um novo cliente"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        dados = request.get_json()
        db = DatabaseConnection()

        query = """
            INSERT INTO clientes
            (Nome_Fantasia, Nome_Razao_Social, CNPJ_CPF, Telefone, Email)
            VALUES (%s, %s, %s, %s, %s)
        """
        params = (
            dados['nome_fantasia'],
            dados['razao_social'],
            dados['cnpj_cpf'],
            dados['telefone'],
            dados['email']
        )

        success = db.execute_update(query, params)

        if success:
            return jsonify({'success': True, 'message': 'Cliente criado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao criar cliente'})

    except Exception as e:
        logging.error(f"❌ Erro ao criar cliente: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/comercial/cliente/editar', methods=['POST'])
def editar_cliente():
    """Edita um cliente existente"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        dados = request.get_json()
        db = DatabaseConnection()

        query = """
            UPDATE clientes
            SET Nome_Fantasia = %s, Nome_Razao_Social = %s, CNPJ_CPF = %s, Telefone = %s, Email = %s
            WHERE id = %s
        """
        params = (
            dados['nome_fantasia'],
            dados['razao_social'],
            dados['cnpj_cpf'],
            dados['telefone'],
            dados['email'],
            dados['id']
        )

        success = db.execute_update(query, params)

        if success:
            return jsonify({'success': True, 'message': 'Cliente atualizado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao atualizar cliente'})

    except Exception as e:
        logging.error(f"❌ Erro ao editar cliente: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/comercial/cliente/excluir/<int:cliente_id>', methods=['DELETE'])
def excluir_cliente(cliente_id):
    """Exclui um cliente"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        db = DatabaseConnection()

        query = "DELETE FROM clientes WHERE id = %s"
        success = db.execute_update(query, (cliente_id,))

        if success:
            return jsonify({'success': True, 'message': 'Cliente excluído com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao excluir cliente'})

    except Exception as e:
        logging.error(f"❌ Erro ao excluir cliente: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

def render_comercial_template(user, metricas, clientes, placas_ativas):
    """Renderiza o template comercial completo"""

    metricas_html = f"""
    <div class="metrics-grid">
        <div class="metric-card">
            <div class="metric-value">{metricas.get('total_clientes', 0)}</div>
            <div class="metric-label">Total de Clientes</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">R$ {metricas.get('faturamento_mensal', 0):,.2f}</div>
            <div class="metric-label">Faturamento Mensal</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{metricas.get('placas_ocupadas', 0)}</div>
            <div class="metric-label">Placas Ocupadas</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">{metricas.get('placas_disponiveis', 0)}</div>
            <div class="metric-label">Placas Disponíveis</div>
        </div>
        <div class="metric-card">
            <div class="metric-value">
                {((metricas.get('placas_ocupadas', 0) / (metricas.get('total_placas', 1)) * 100) if metricas.get('total_placas', 0) > 0 else 0):.1f}%
            </div>
            <div class="metric-label">Taxa de Ocupação</div>
        </div>
    </div>
    """

    clientes_html = ""
    for cliente in clientes:
        clientes_html += f"""
        <div class="cliente-card">
            <div class="cliente-header">
                <h4>{cliente['Nome_Fantasia']}</h4>
                <div class="cliente-actions">
                    <button class="btn-editar-cliente" onclick="editarCliente({cliente['id']})">✏️</button>
                    <button class="btn-excluir-cliente" onclick="excluirCliente({cliente['id']})">🗑️</button>
                </div>
            </div>
            <div class="cliente-info">
                <p><strong>Razão Social:</strong> {cliente['Nome_Razao_Social']}</p>
                <p><strong>CNPJ/CPF:</strong> {cliente['CNPJ_CPF']}</p>
                <p><strong>Telefone:</strong> {cliente['Telefone']}</p>
                <p><strong>Email:</strong> {cliente['Email']}</p>
            </div>
        </div>
        """

    # Converter clientes para JSON serializable
    def convert_to_serializable(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, datetime.date):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {key: convert_to_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [convert_to_serializable(item) for item in obj]
        else:
            return obj

    clientes_serializable = convert_to_serializable(clientes)

    return f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Comercial - SIG-ME</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: {CORES_FORMATO['claro']};
            color: {CORES_FORMATO['texto']};
        }}
        .header {{
            background: {CORES_FORMATO['branco']};
            padding: 20px 30px;
            box-shadow: 0 2px 20px rgba(0, 0, 0, 0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 4px solid {CORES_FORMATO['destaque']};
        }}
        .logo h1 {{
            color: {CORES_FORMATO['primaria']};
            font-size: 24px;
            font-weight: 700;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
            color: {CORES_FORMATO['texto']};
            font-weight: 500;
        }}
        .btn-logout {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
        }}
        .btn-logout:hover {{
            background: {CORES_FORMATO['alerta']};
            transform: translateY(-2px);
        }}
        .container {{
            max-width: 1400px;
            margin: 30px auto;
            padding: 0 25px;
        }}
        .navigation {{
            display: flex;
            gap: 12px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .nav-btn {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 14px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .nav-btn:hover {{
            background: {CORES_FORMATO['secundaria']};
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(26, 54, 93, 0.2);
        }}
        .nav-btn.active {{
            background: {CORES_FORMATO['destaque']};
        }}
        .content {{
            background: {CORES_FORMATO['branco']};
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        }}
        .section-title {{
            font-size: 24px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .metric-card {{
            background: {CORES_FORMATO['claro']};
            padding: 25px;
            border-radius: 10px;
            text-align: center;
            border-left: 4px solid {CORES_FORMATO['sucesso']};
        }}
        .metric-value {{
            font-size: 32px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 8px;
        }}
        .metric-label {{
            color: {CORES_FORMATO['texto_claro']};
            font-size: 14px;
            font-weight: 600;
        }}
        .card {{
            background: {CORES_FORMATO['claro']};
            padding: 25px;
            border-radius: 10px;
            margin-top: 30px;
        }}
        .card h3 {{
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            font-size: 20px;
        }}
        .clientes-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
            gap: 20px;
            margin-top: 15px;
        }}
        .cliente-card {{
            background: {CORES_FORMATO['branco']};
            padding: 20px;
            border-radius: 8px;
            border: 1px solid {CORES_FORMATO['claro']};
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }}
        .cliente-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
        }}
        .cliente-actions {{
            display: flex;
            gap: 5px;
        }}
        .btn-editar-cliente {{
            background: {CORES_FORMATO['info']};
            color: {CORES_FORMATO['branco']};
            border: none;
            padding: 5px 10px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
        }}
        .btn-excluir-cliente {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            border: none;
            padding: 5px 10px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
        }}
        .cliente-card h4 {{
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 10px;
        }}
        .cliente-info p {{
            margin: 5px 0;
            font-size: 14px;
        }}
        .btn-novo-cliente {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            margin-bottom: 20px;
            transition: all 0.3s;
        }}
        .btn-novo-cliente:hover {{
            background: {CORES_FORMATO['info']};
            transform: translateY(-2px);
        }}
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow-y: auto;
        }}
        .modal-content {{
            background-color: {CORES_FORMATO['branco']};
            margin: 5% auto;
            padding: 30px;
            border-radius: 10px;
            width: 90%;
            max-width: 500px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.2);
            max-height: 85vh;
            overflow-y: auto;
        }}
        .modal-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .close {{
            color: {CORES_FORMATO['texto_claro']};
            font-size: 28px;
            font-weight: bold;
            cursor: pointer;
        }}
        .close:hover {{
            color: {CORES_FORMATO['destaque']};
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-group label {{
            display: block;
            margin-bottom: 5px;
            font-weight: 600;
            color: {CORES_FORMATO['primaria']};
        }}
        .form-group input {{
            width: 100%;
            padding: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
            border-radius: 5px;
            font-size: 14px;
        }}
        .form-actions {{
            display: flex;
            gap: 10px;
            justify-content: flex-end;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid {CORES_FORMATO['claro']};
        }}
        .btn-cancelar {{
            background: {CORES_FORMATO['texto_claro']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: 600;
        }}
        .btn-salvar {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: 600;
        }}
        .btn-cancelar:hover {{
            background: {CORES_FORMATO['secundaria']};
        }}
        .btn-salvar:hover {{
            background: #2f855a;
        }}
        .alert {{
            padding: 15px;
            margin: 15px 0;
            border-radius: 5px;
            font-weight: 500;
        }}
        .alert.success {{
            background: {CORES_FORMATO['sucesso']}20;
            color: {CORES_FORMATO['sucesso']};
            border: 1px solid {CORES_FORMATO['sucesso']};
        }}
        .alert.error {{
            background: {CORES_FORMATO['destaque']}20;
            color: {CORES_FORMATO['destaque']};
            border: 1px solid {CORES_FORMATO['destaque']};
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <h1>🚀 SIG-ME - Comercial</h1>
        </div>
        <div class="user-info">
            <span>👋 Olá, {user['nome']} <small>({user['perfil']})</small></span>
            <a href="/logout" class="btn-logout">🚪 Sair</a>
        </div>
    </div>

    <div class="container">
        <div class="navigation">
            <a href="/inventario" class="nav-btn">📋 Inventário</a>
            <a href="/relatorios" class="nav-btn">📊 Relatórios</a>
            <a href="/comercial" class="nav-btn active">💼 Comercial</a>
            <a href="/contratos" class="nav-btn">📑 Contratos</a>
            <a href="/exportar_dados" class="nav-btn">📥 Exportar Dados</a>
        </div>

        <div class="content">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                <h2 class="section-title">💼 Módulo Comercial</h2>
                <button class="btn-novo-cliente" onclick="abrirModalNovoCliente()">➕ Novo Cliente</button>
            </div>
            <div id="alert-container"></div>

            {metricas_html}

            <div class="card">
                <h3>👥 Clientes Cadastrados</h3>
                <div class="clientes-grid">
                    {clientes_html if clientes_html else '<p>Nenhum cliente cadastrado</p>'}
                </div>
            </div>
        </div>
    </div>

    <!-- Modal Cliente -->
    <div id="modalCliente" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h3 id="modalTituloCliente">Novo Cliente</h3>
                <span class="close" onclick="fecharModalCliente()">&times;</span>
            </div>
            <form id="formCliente">
                <input type="hidden" id="cliente_id" name="cliente_id">
                <div class="form-group">
                    <label for="nome_fantasia">Nome Fantasia:</label>
                    <input type="text" id="nome_fantasia" name="nome_fantasia" required>
                </div>
                <div class="form-group">
                    <label for="razao_social">Razão Social:</label>
                    <input type="text" id="razao_social" name="razao_social" required>
                </div>
                <div class="form-group">
                    <label for="cnpj_cpf">CNPJ/CPF:</label>
                    <input type="text" id="cnpj_cpf" name="cnpj_cpf" required>
                </div>
                <div class="form-group">
                    <label for="telefone">Telefone:</label>
                    <input type="text" id="telefone" name="telefone" required>
                </div>
                <div class="form-group">
                    <label for="email">Email:</label>
                    <input type="email" id="email" name="email" required>
                </div>
                <div class="form-actions">
                    <button type="button" class="btn-cancelar" onclick="fecharModalCliente()">Cancelar</button>
                    <button type="submit" class="btn-salvar">💾 Salvar Cliente</button>
                </div>
            </form>
        </div>
    </div>
    <script>
        let clienteEditando = null;
        const clientes = {json.dumps(clientes_serializable)};

        function abrirModalNovoCliente() {{
            document.getElementById('modalTituloCliente').textContent = 'Novo Cliente';
            document.getElementById('formCliente').reset();
            document.getElementById('cliente_id').value = '';
            clienteEditando = null;
            document.getElementById('modalCliente').style.display = 'block';
        }}

        function fecharModalCliente() {{
            document.getElementById('modalCliente').style.display = 'none';
        }}

        function editarCliente(clienteId) {{
            const cliente = clientes.find(c => c.id === clienteId);

            if (cliente) {{
                document.getElementById('modalTituloCliente').textContent = 'Editar Cliente';
                document.getElementById('cliente_id').value = cliente.id;
                document.getElementById('nome_fantasia').value = cliente.Nome_Fantasia;
                document.getElementById('razao_social').value = cliente.Nome_Razao_Social;
                document.getElementById('cnpj_cpf').value = cliente.CNPJ_CPF;
                document.getElementById('telefone').value = cliente.Telefone;
                document.getElementById('email').value = cliente.Email;

                clienteEditando = cliente;
                document.getElementById('modalCliente').style.display = 'block';
            }}
        }}

        function excluirCliente(clienteId) {{
            if (confirm('Tem certeza que deseja excluir este cliente?')) {{
                fetch('/comercial/cliente/excluir/' + clienteId, {{ method: 'DELETE' }})
                    .then(response => response.json())
                    .then(data => {{
                        if (data.success) {{
                            mostrarAlert(data.message, 'success');
                            setTimeout(() => location.reload(), 1500);
                        }} else {{
                            mostrarAlert(data.message, 'error');
                        }}
                    }})
                    .catch(error => {{
                        console.error('Erro:', error);
                        mostrarAlert('Erro ao excluir cliente', 'error');
                    }});
            }}
        }}

        function mostrarAlert(mensagem, tipo) {{
            const alertContainer = document.getElementById('alert-container');
            const alert = document.createElement('div');
            alert.className = 'alert ' + tipo;
            alert.textContent = mensagem;
            alertContainer.appendChild(alert);

            setTimeout(() => {{
                alert.remove();
            }}, 5000);
        }}

        document.getElementById('formCliente').addEventListener('submit', function(e) {{
            e.preventDefault();

            const formData = new FormData(this);
            const clienteId = formData.get('cliente_id');

            const dados = {{
                nome_fantasia: formData.get('nome_fantasia'),
                razao_social: formData.get('razao_social'),
                cnpj_cpf: formData.get('cnpj_cpf'),
                telefone: formData.get('telefone'),
                email: formData.get('email')
            }};

            if (clienteId) {{
                // Editar cliente existente
                dados.id = parseInt(clienteId);
                fetch('/comercial/cliente/editar', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify(dados)
                }})
                .then(response => response.json())
                .then(data => {{
                    if (data.success) {{
                        mostrarAlert(data.message, 'success');
                        fecharModalCliente();
                        setTimeout(() => location.reload(), 1500);
                    }} else {{
                        mostrarAlert(data.message, 'error');
                    }}
                }})
                .catch(error => {{
                    console.error('Erro:', error);
                    mostrarAlert('Erro ao editar cliente', 'error');
                }});
            }} else {{
                // Novo cliente
                fetch('/comercial/cliente/novo', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify(dados)
                }})
                .then(response => response.json())
                .then(data => {{
                    if (data.success) {{
                        mostrarAlert(data.message, 'success');
                        fecharModalCliente();
                        setTimeout(() => location.reload(), 1500);
                    }} else {{
                        mostrarAlert(data.message, 'error');
                    }}
                }})
                .catch(error => {{
                    console.error('Erro:', error);
                    mostrarAlert('Erro ao criar cliente', 'error');
                }});
            }}
        }});

        // Fechar modal ao clicar fora
        window.onclick = function(event) {{
            const modal = document.getElementById('modalCliente');
            if (event.target === modal) {{
                fecharModalCliente();
            }}
        }}
    </script>
</body>
</html>
'''

# =============================================================================
# MÓDULO DE CONTRATOS (COMPLETAMENTE CORRIGIDO)
# =============================================================================
@app.route('/contratos')
def contratos():
    """Página de contratos - COMPLETAMENTE CORRIGIDA"""
    if 'user' not in session:
        return redirect('/login')

    db = DatabaseConnection()

    # Buscar contratos com informações completas
    contratos = db.execute_query("""
        SELECT c.*, cl.Nome_Fantasia as cliente_nome, f.razao_social as fornecedor_nome
        FROM contratos c
        LEFT JOIN clientes cl ON c.cliente_id = cl.id
        LEFT JOIN fornecedores f ON c.fornecedor_id = f.id
        ORDER BY c.data_inicio DESC
    """) or []

    # Buscar clientes e fornecedores para os dropdowns
    clientes = db.execute_query("SELECT id, Nome_Fantasia FROM clientes ORDER BY Nome_Fantasia") or []
    fornecedores = db.execute_query("SELECT id, razao_social FROM fornecedores ORDER BY razao_social") or []

    return render_contratos_template(session['user'], contratos, clientes, fornecedores)

@app.route('/contratos/novo', methods=['POST'])
def novo_contrato():
    """Cria um novo contrato - CORRIGIDO"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        dados = request.get_json()
        db = DatabaseConnection()

        # Determinar IDs baseado no tipo
        cliente_id = dados.get('cliente_id') if dados['tipo'] == 'cliente' else None
        fornecedor_id = dados.get('fornecedor_id') if dados['tipo'] == 'fornecedor' else None

        query = """
            INSERT INTO contratos
            (tipo, cliente_id, fornecedor_id, descricao, valor_mensal, data_inicio, data_fim, status, observacoes, arquivo_pdf)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        params = (
            dados['tipo'],
            cliente_id,
            fornecedor_id,
            dados['descricao'],
            dados['valor_mensal'],
            dados['data_inicio'],
            dados.get('data_fim'),
            dados.get('status', 'ativo'),
            dados.get('observacoes', ''),
            dados.get('arquivo_pdf', '')
        )

        success = db.execute_update(query, params)

        if success:
            return jsonify({'success': True, 'message': 'Contrato criado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao criar contrato'})

    except Exception as e:
        logging.error(f"❌ Erro ao criar contrato: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/contratos/editar', methods=['POST'])
def editar_contrato():
    """Edita um contrato existente - NOVA FUNÇÃO"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        dados = request.get_json()
        db = DatabaseConnection()

        # Determinar IDs baseado no tipo
        cliente_id = dados.get('cliente_id') if dados['tipo'] == 'cliente' else None
        fornecedor_id = dados.get('fornecedor_id') if dados['tipo'] == 'fornecedor' else None

        query = """
            UPDATE contratos
            SET tipo = %s, cliente_id = %s, fornecedor_id = %s, descricao = %s,
                valor_mensal = %s, data_inicio = %s, data_fim = %s, status = %s,
                observacoes = %s, arquivo_pdf = %s
            WHERE id = %s
        """
        params = (
            dados['tipo'],
            cliente_id,
            fornecedor_id,
            dados['descricao'],
            dados['valor_mensal'],
            dados['data_inicio'],
            dados.get('data_fim'),
            dados.get('status', 'ativo'),
            dados.get('observacoes', ''),
            dados.get('arquivo_pdf', ''),
            dados['id']
        )

        success = db.execute_update(query, params)

        if success:
            return jsonify({'success': True, 'message': 'Contrato atualizado com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao atualizar contrato'})

    except Exception as e:
        logging.error(f"❌ Erro ao editar contrato: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

@app.route('/contratos/excluir/<int:contrato_id>', methods=['DELETE'])
def excluir_contrato(contrato_id):
    """Exclui um contrato"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Não autorizado'})

    try:
        db = DatabaseConnection()

        query = "DELETE FROM contratos WHERE id = %s"
        success = db.execute_update(query, (contrato_id,))

        if success:
            return jsonify({'success': True, 'message': 'Contrato excluído com sucesso!'})
        else:
            return jsonify({'success': False, 'message': 'Erro ao excluir contrato'})

    except Exception as e:
        logging.error(f"❌ Erro ao excluir contrato: {e}")
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'})

def render_contratos_template(user, contratos, clientes, fornecedores):
    """Renderiza o template de contratos COMPLETAMENTE CORRIGIDO"""

    contratos_html = ""
    for contrato in contratos:
        # Determinar nome do contratante
        if contrato['tipo'] == 'cliente':
            nome_contratante = contrato.get('cliente_nome', 'Cliente não especificado')
            tipo_contratante = 'Cliente'
        else:
            nome_contratante = contrato.get('fornecedor_nome', 'Fornecedor não especificado')
            tipo_contratante = 'Fornecedor'

        data_fim = contrato.get('data_fim') or 'Não definida'
        arquivo_pdf = contrato.get('arquivo_pdf') or 'Nenhum arquivo'
        descricao = contrato.get('descricao', 'Sem descrição')
        valor_mensal = contrato.get('valor_mensal', 0)
        data_inicio = contrato.get('data_inicio', 'Não definida')
        status = contrato.get('status', 'ativo')
        observacoes = contrato.get('observacoes', '')

        contratos_html += f"""
        <div class="contrato-card">
            <div class="contrato-header">
                <h4>📋 {descricao[:50]}{'...' if len(descricao) > 50 else ''}</h4>
                <div class="contrato-actions">
                    <button class="btn-editar" onclick="editarContrato({contrato.get('id', 0)})">✏️ Editar</button>
                    <button class="btn-pdf" onclick="visualizarPDF({contrato.get('id', 0)})">📄 PDF</button>
                    <button class="btn-excluir" onclick="excluirContrato({contrato.get('id', 0)})">🗑️ Excluir</button>
                </div>
            </div>
            <div class="contrato-info">
                <p><strong>{tipo_contratante}:</strong> {nome_contratante}</p>
                <p><strong>Valor Mensal:</strong> R$ {valor_mensal:,.2f}</p>
                <p><strong>Início:</strong> {data_inicio}</p>
                <p><strong>Término:</strong> {data_fim}</p>
                <p><strong>Status:</strong> <span class="status-badge {status}">{status}</span></p>
                <p><strong>Arquivo:</strong> {arquivo_pdf}</p>
                <p><strong>Observações:</strong> {observacoes[:100]}{'...' if len(observacoes) > 100 else ''}</p>
            </div>
        </div>
        """

    # Converter para JSON serializable
    def convert_to_serializable(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, datetime.date):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {key: convert_to_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [convert_to_serializable(item) for item in obj]
        else:
            return obj

    contratos_serializable = convert_to_serializable(contratos)
    clientes_serializable = convert_to_serializable(clientes)
    fornecedores_serializable = convert_to_serializable(fornecedores)

    return f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Contratos - SIG-ME</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: {CORES_FORMATO['claro']};
            color: {CORES_FORMATO['texto']};
        }}
        .header {{
            background: {CORES_FORMATO['branco']};
            padding: 20px 30px;
            box-shadow: 0 2px 20px rgba(0, 0, 0, 0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 4px solid {CORES_FORMATO['destaque']};
        }}
        .logo h1 {{
            color: {CORES_FORMATO['primaria']};
            font-size: 24px;
            font-weight: 700;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
            color: {CORES_FORMATO['texto']};
            font-weight: 500;
        }}
        .btn-logout {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
        }}
        .btn-logout:hover {{
            background: {CORES_FORMATO['alerta']};
            transform: translateY(-2px);
        }}
        .container {{
            max-width: 1400px;
            margin: 30px auto;
            padding: 0 25px;
        }}
        .navigation {{
            display: flex;
            gap: 12px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .nav-btn {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 14px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .nav-btn:hover {{
            background: {CORES_FORMATO['secundaria']};
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(26, 54, 93, 0.2);
        }}
        .nav-btn.active {{
            background: {CORES_FORMATO['destaque']};
        }}
        .content {{
            background: {CORES_FORMATO['branco']};
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        }}
        .section-title {{
            font-size: 24px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .contratos-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(400px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }}
        .contrato-card {{
            background: {CORES_FORMATO['branco']};
            padding: 20px;
            border-radius: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
        }}
        .contrato-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 1px solid {CORES_FORMATO['claro']};
        }}
        .contrato-actions {{
            display: flex;
            gap: 8px;
        }}
        .btn-editar {{
            background: {CORES_FORMATO['info']};
            color: {CORES_FORMATO['branco']};
            border: none;
            padding: 8px 12px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
        }}
        .btn-pdf {{
            background: {CORES_FORMATO['alerta']};
            color: {CORES_FORMATO['branco']};
            border: none;
            padding: 8px 12px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
        }}
        .btn-excluir {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            border: none;
            padding: 8px 12px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
        }}
        .contrato-info p {{
            margin: 8px 0;
            font-size: 14px;
        }}
        .status-badge {{
            padding: 4px 8px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }}
        .status-badge.ativo {{
            background: {CORES_FORMATO['sucesso']}20;
            color: {CORES_FORMATO['sucesso']};
        }}
        .status-badge.inativo {{
            background: {CORES_FORMATO['texto_claro']}20;
            color: {CORES_FORMATO['texto_claro']};
        }}
        .status-badge.vencido {{
            background: {CORES_FORMATO['destaque']}20;
            color: {CORES_FORMATO['destaque']};
        }}
        .btn-novo-contrato {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            margin-bottom: 20px;
            transition: all 0.3s;
        }}
        .btn-novo-contrato:hover {{
            background: {CORES_FORMATO['info']};
            transform: translateY(-2px);
        }}
        .alert {{
            padding: 15px;
            margin: 15px 0;
            border-radius: 5px;
            font-weight: 500;
        }}
        .alert.success {{
            background: {CORES_FORMATO['sucesso']}20;
            color: {CORES_FORMATO['sucesso']};
            border: 1px solid {CORES_FORMATO['sucesso']};
        }}
        .alert.error {{
            background: {CORES_FORMATO['destaque']}20;
            color: {CORES_FORMATO['destaque']};
            border: 1px solid {CORES_FORMATO['destaque']};
        }}
        /* MODAL STYLES - ATUALIZADO COM SCROLL */
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
            overflow-y: auto;
        }}
        .modal-content {{
            background-color: {CORES_FORMATO['branco']};
            margin: 5% auto;
            padding: 30px;
            border-radius: 10px;
            width: 90%;
            max-width: 600px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.2);
            max-height: 85vh;
            overflow-y: auto;
        }}
        .modal-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            padding-bottom: 15px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .close {{
            color: {CORES_FORMATO['texto_claro']};
            font-size: 28px;
            font-weight: bold;
            cursor: pointer;
        }}
        .close:hover {{
            color: {CORES_FORMATO['destaque']};
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-group label {{
            display: block;
            margin-bottom: 5px;
            font-weight: 600;
            color: {CORES_FORMATO['primaria']};
        }}
        .form-group input, .form-group select, .form-group textarea {{
            width: 100%;
            padding: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
            border-radius: 5px;
            font-size: 14px;
        }}
        .form-group textarea {{
            height: 100px;
            resize: vertical;
        }}
        .form-actions {{
            display: flex;
            gap: 10px;
            justify-content: flex-end;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid {CORES_FORMATO['claro']};
        }}
        .btn-cancelar {{
            background: {CORES_FORMATO['texto_claro']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: 600;
        }}
        .btn-salvar {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: 600;
        }}
        .btn-cancelar:hover {{
            background: {CORES_FORMATO['secundaria']};
        }}
        .btn-salvar:hover {{
            background: #2f855a;
        }}
        .file-upload {{
            border: 2px dashed {CORES_FORMATO['claro']};
            padding: 20px;
            text-align: center;
            border-radius: 5px;
            cursor: pointer;
        }}
        .file-upload:hover {{
            border-color: {CORES_FORMATO['info']};
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <h1>🚀 SIG-ME - Contratos</h1>
        </div>
        <div class="user-info">
            <span>👋 Olá, {user['nome']} <small>({user['perfil']})</small></span>
            <a href="/logout" class="btn-logout">🚪 Sair</a>
        </div>
    </div>

    <div class="container">
        <div class="navigation">
            <a href="/inventario" class="nav-btn">📋 Inventário</a>
            <a href="/relatorios" class="nav-btn">📊 Relatórios</a>
            <a href="/comercial" class="nav-btn">💼 Comercial</a>
            <a href="/contratos" class="nav-btn active">📑 Contratos</a>
            <a href="/exportar_dados" class="nav-btn">📥 Exportar Dados</a>
        </div>

        <div class="content">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
                <h2 class="section-title">📑 Módulo de Contratos</h2>
                <button class="btn-novo-contrato" onclick="abrirModalNovoContrato()">➕ Novo Contrato</button>
            </div>
            <div id="alert-container"></div>

            <div class="contratos-grid">
                {contratos_html if contratos_html else '<p>Nenhum contrato cadastrado</p>'}
            </div>
        </div>
    </div>

    <!-- Modal Novo/Editar Contrato - ATUALIZADO COM SCROLL -->
    <div id="modalContrato" class="modal">
        <div class="modal-content">
            <div class="modal-header">
                <h3 id="modalTituloContrato">Novo Contrato</h3>
                <span class="close" onclick="fecharModalContrato()">&times;</span>
            </div>
            <form id="formContrato">
                <input type="hidden" id="contrato_id" name="contrato_id">

                <div class="form-group">
                    <label for="tipo_contrato">Tipo de Contrato:</label>
                    <select id="tipo_contrato" name="tipo_contrato" required onchange="toggleContratante()">
                        <option value="cliente">Cliente</option>
                        <option value="fornecedor">Fornecedor</option>
                    </select>
                </div>

                <div class="form-group" id="cliente-group">
                    <label for="cliente_id">Cliente:</label>
                    <select id="cliente_id" name="cliente_id">
                        <option value="">Selecione um cliente</option>
                        {"".join([f'<option value="{cliente["id"]}">{cliente["Nome_Fantasia"]}</option>' for cliente in clientes])}
                    </select>
                </div>

                <div class="form-group" id="fornecedor-group" style="display: none;">
                    <label for="fornecedor_id">Fornecedor:</label>
                    <select id="fornecedor_id" name="fornecedor_id">
                        <option value="">Selecione um fornecedor</option>
                        {"".join([f'<option value="{fornecedor["id"]}">{fornecedor["razao_social"]}</option>' for fornecedor in fornecedores])}
                    </select>
                </div>

                <div class="form-group">
                    <label for="descricao">Descrição do Contrato:</label>
                    <textarea id="descricao" name="descricao" required placeholder="Descreva o contrato..."></textarea>
                </div>

                <div class="form-group">
                    <label for="valor_mensal">Valor Mensal (R$):</label>
                    <input type="number" id="valor_mensal" name="valor_mensal" step="0.01" min="0" required>
                </div>

                <div class="form-group">
                    <label for="data_inicio">Data de Início:</label>
                    <input type="date" id="data_inicio" name="data_inicio" required>
                </div>

                <div class="form-group">
                    <label for="data_fim">Data de Término (opcional):</label>
                    <input type="date" id="data_fim" name="data_fim">
                </div>

                <div class="form-group">
                    <label for="status">Status:</label>
                    <select id="status" name="status" required>
                        <option value="ativo">Ativo</option>
                        <option value="inativo">Inativo</option>
                        <option value="vencido">Vencido</option>
                    </select>
                </div>

                <div class="form-group">
                    <label for="observacoes">Observações:</label>
                    <textarea id="observacoes" name="observacoes" placeholder="Observações adicionais..."></textarea>
                </div>

                <div class="form-group">
                    <label for="arquivo_pdf">Arquivo PDF (nome do arquivo):</label>
                    <input type="text" id="arquivo_pdf" name="arquivo_pdf" placeholder="ex: contrato_cliente_001.pdf">
                </div>

                <div class="form-actions">
                    <button type="button" class="btn-cancelar" onclick="fecharModalContrato()">Cancelar</button>
                    <button type="submit" class="btn-salvar">💾 Salvar Contrato</button>
                </div>
            </form>
        </div>
    </div>

    <script>
        let contratoEditando = null;
        const contratos = {json.dumps(contratos_serializable)};
        const clientes = {json.dumps(clientes_serializable)};
        const fornecedores = {json.dumps(fornecedores_serializable)};

        function abrirModalNovoContrato() {{
            document.getElementById('modalTituloContrato').textContent = 'Novo Contrato';
            document.getElementById('formContrato').reset();
            document.getElementById('contrato_id').value = '';
            document.getElementById('data_inicio').valueAsDate = new Date();
            contratoEditando = null;
            document.getElementById('modalContrato').style.display = 'block';
        }}

        function fecharModalContrato() {{
            document.getElementById('modalContrato').style.display = 'none';
        }}

        function toggleContratante() {{
            const tipo = document.getElementById('tipo_contrato').value;
            if (tipo === 'cliente') {{
                document.getElementById('cliente-group').style.display = 'block';
                document.getElementById('fornecedor-group').style.display = 'none';
            }} else {{
                document.getElementById('cliente-group').style.display = 'none';
                document.getElementById('fornecedor-group').style.display = 'block';
            }}
        }}

        function editarContrato(contratoId) {{
            const contrato = contratos.find(c => c.id === contratoId);

            if (contrato) {{
                document.getElementById('modalTituloContrato').textContent = 'Editar Contrato';
                document.getElementById('contrato_id').value = contrato.id;
                document.getElementById('tipo_contrato').value = contrato.tipo;
                document.getElementById('descricao').value = contrato.descricao;
                document.getElementById('valor_mensal').value = contrato.valor_mensal;
                document.getElementById('data_inicio').value = contrato.data_inicio;
                document.getElementById('data_fim').value = contrato.data_fim || '';
                document.getElementById('status').value = contrato.status;
                document.getElementById('observacoes').value = contrato.observacoes || '';
                document.getElementById('arquivo_pdf').value = contrato.arquivo_pdf || '';

                if (contrato.tipo === 'cliente') {{
                    document.getElementById('cliente_id').value = contrato.cliente_id || '';
                }} else {{
                    document.getElementById('fornecedor_id').value = contrato.fornecedor_id || '';
                }}

                toggleContratante();
                contratoEditando = contrato;
                document.getElementById('modalContrato').style.display = 'block';
            }}
        }}

        function visualizarPDF(contratoId) {{
            // Abre o PDF em uma nova aba
            window.open('/contratos/pdf/' + contratoId, '_blank');
        }}

        function excluirContrato(contratoId) {{
            if (confirm('Tem certeza que deseja excluir este contrato?')) {{
                fetch('/contratos/excluir/' + contratoId, {{ method: 'DELETE' }})
                    .then(response => response.json())
                    .then(data => {{
                        if (data.success) {{
                            mostrarAlert(data.message, 'success');
                            setTimeout(() => location.reload(), 1500);
                        }} else {{
                            mostrarAlert(data.message, 'error');
                        }}
                    }})
                    .catch(error => {{
                        console.error('Erro:', error);
                        mostrarAlert('Erro ao excluir contrato', 'error');
                    }});
            }}
        }}

        function mostrarAlert(mensagem, tipo) {{
            const alertContainer = document.getElementById('alert-container');
            const alert = document.createElement('div');
            alert.className = 'alert ' + tipo;
            alert.textContent = mensagem;
            alertContainer.appendChild(alert);

            setTimeout(() => {{
                alert.remove();
            }}, 5000);
        }}

        // Envio do formulário
        document.getElementById('formContrato').addEventListener('submit', function(e) {{
            e.preventDefault();

            const formData = new FormData(this);
            const contratoId = document.getElementById('contrato_id').value;

            const dados = {{
                tipo: document.getElementById('tipo_contrato').value,
                descricao: document.getElementById('descricao').value,
                valor_mensal: parseFloat(document.getElementById('valor_mensal').value),
                data_inicio: document.getElementById('data_inicio').value,
                data_fim: document.getElementById('data_fim').value || null,
                status: document.getElementById('status').value,
                observacoes: document.getElementById('observacoes').value,
                arquivo_pdf: document.getElementById('arquivo_pdf').value || ''
            }};

            // Adicionar cliente ou fornecedor
            if (dados.tipo === 'cliente') {{
                dados.cliente_id = document.getElementById('cliente_id').value || null;
                dados.fornecedor_id = null;
            }} else {{
                dados.fornecedor_id = document.getElementById('fornecedor_id').value || null;
                dados.cliente_id = null;
            }}

            // Adicionar ID se estiver editando
            if (contratoId) {{
                dados.id = parseInt(contratoId);
            }}

            const url = contratoId ? '/contratos/editar' : '/contratos/novo';

            fetch(url, {{
                method: 'POST',
                headers: {{
                    'Content-Type': 'application/json',
                }},
                body: JSON.stringify(dados)
            }})
            .then(response => response.json())
            .then(data => {{
                if (data.success) {{
                    mostrarAlert(data.message, 'success');
                    fecharModalContrato();
                    setTimeout(() => location.reload(), 1500);
                }} else {{
                    mostrarAlert(data.message, 'error');
                }}
            }})
            .catch(error => {{
                console.error('Erro:', error);
                mostrarAlert('Erro ao salvar contrato', 'error');
            }});
        }});

        // Fechar modal ao clicar fora
        window.onclick = function(event) {{
            const modal = document.getElementById('modalContrato');
            if (event.target === modal) {{
                fecharModalContrato();
            }}
        }}

        // Inicializar
        toggleContratante();
    </script>
</body>
</html>
'''

# =============================================================================
# MÓDULO DE RELATÓRIOS (CORRIGIDO COM EXPORTAÇÃO)
# =============================================================================
@app.route('/relatorios')
def relatorios():
    """Página de relatórios - CORRIGIDA COM EXPORTAÇÃO"""
    if 'user' not in session:
        return redirect('/login')

    db = DatabaseConnection()

    # Gerar relatórios
    relatorio_ocupacao = relatorio_manager.gerar_relatorio_ocupacao()
    relatorio_financeiro = relatorio_manager.gerar_relatorio_financeiro()
    relatorio_regiao = relatorio_manager.gerar_relatorio_regiao()

    return f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatórios - SIG-ME</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: {CORES_FORMATO['claro']};
            color: {CORES_FORMATO['texto']};
        }}
        .header {{
            background: {CORES_FORMATO['branco']};
            padding: 20px 30px;
            box-shadow: 0 2px 20px rgba(0, 0, 0, 0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 4px solid {CORES_FORMATO['destaque']};
        }}
        .logo h1 {{
            color: {CORES_FORMATO['primaria']};
            font-size: 24px;
            font-weight: 700;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
            color: {CORES_FORMATO['texto']};
            font-weight: 500;
        }}
        .btn-logout {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
        }}
        .btn-logout:hover {{
            background: {CORES_FORMATO['alerta']};
            transform: translateY(-2px);
        }}
        .container {{
            max-width: 1400px;
            margin: 30px auto;
            padding: 0 25px;
        }}
        .navigation {{
            display: flex;
            gap: 12px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .nav-btn {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 14px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .nav-btn:hover {{
            background: {CORES_FORMATO['secundaria']};
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(26, 54, 93, 0.2);
        }}
        .nav-btn.active {{
            background: {CORES_FORMATO['destaque']};
        }}
        .content {{
            background: {CORES_FORMATO['branco']};
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        }}
        .section-title {{
            font-size: 24px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .relatorios-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 25px;
            margin-top: 20px;
        }}
        .relatorio-card {{
            background: {CORES_FORMATO['claro']};
            padding: 25px;
            border-radius: 10px;
            border: 1px solid {CORES_FORMATO['claro']};
        }}
        .relatorio-card h3 {{
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 15px;
            font-size: 18px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .relatorio-item {{
            background: {CORES_FORMATO['branco']};
            padding: 15px;
            border-radius: 6px;
            margin-bottom: 10px;
        }}
        .relatorio-item p {{
            margin: 5px 0;
            font-size: 14px;
        }}
        .btn-exportar {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 8px 16px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 12px;
            text-decoration: none;
            display: inline-block;
        }}
        .btn-exportar:hover {{
            background: {CORES_FORMATO['info']};
        }}
        .filtros {{
            background: {CORES_FORMATO['claro']};
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .filtros h4 {{
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 10px;
        }}
        .filtro-group {{
            display: flex;
            gap: 10px;
            align-items: center;
        }}
        .filtro-group input {{
            padding: 8px;
            border: 1px solid {CORES_FORMATO['claro']};
            border-radius: 4px;
        }}
        .btn-filtrar {{
            background: {CORES_FORMATO['info']};
            color: {CORES_FORMATO['branco']};
            padding: 8px 16px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <h1>🚀 SIG-ME - Relatórios</h1>
        </div>
        <div class="user-info">
            <span>👋 Olá, {session['user']['nome']} <small>({session['user']['perfil']})</small></span>
            <a href="/logout" class="btn-logout">🚪 Sair</a>
        </div>
    </div>

    <div class="container">
        <div class="navigation">
            <a href="/inventario" class="nav-btn">📋 Inventário</a>
            <a href="/relatorios" class="nav-btn active">📊 Relatórios</a>
            <a href="/comercial" class="nav-btn">💼 Comercial</a>
            <a href="/contratos" class="nav-btn">📑 Contratos</a>
            <a href="/exportar_dados" class="nav-btn">📥 Exportar Dados</a>
        </div>

        <div class="content">
            <h2 class="section-title">📊 Módulo de Relatórios</h2>

            <div class="filtros">
                <h4>🔍 Filtros por Data</h4>
                <div class="filtro-group">
                    <label>Data Início:</label>
                    <input type="date" id="data_inicio">
                    <label>Data Fim:</label>
                    <input type="date" id="data_fim">
                    <button class="btn-filtrar" onclick="aplicarFiltros()">Aplicar Filtros</button>
                    <button class="btn-filtrar" onclick="limparFiltros()" style="background: {CORES_FORMATO['texto_claro']};">Limpar</button>
                </div>
            </div>

            <div class="relatorios-grid">
                <div class="relatorio-card">
                    <h3>📊 Ocupação por Região
                        <a href="/relatorios/exportar/ocupacao" class="btn-exportar">📥 Exportar Excel</a>
                    </h3>
                    {generate_ocupacao_html(relatorio_ocupacao)}
                </div>

                <div class="relatorio-card">
                    <h3>💰 Relatórios Financeiros
                        <a href="/relatorios/exportar/financeiro" class="btn-exportar">📥 Exportar Excel</a>
                    </h3>
                    {generate_financeiro_html(relatorio_financeiro)}
                </div>

                <div class="relatorio-card">
                    <h3>🏙️ Resumo por Região
                        <a href="/relatorios/exportar/inventario" class="btn-exportar">📥 Exportar Excel</a>
                    </h3>
                    {generate_regiao_html(relatorio_regiao)}
                </div>
            </div>
        </div>
    </div>

    <script>
        function aplicarFiltros() {{
            const dataInicio = document.getElementById('data_inicio').value;
            const dataFim = document.getElementById('data_fim').value;

            if (dataInicio || dataFim) {{
                alert('Filtros aplicados!\\nData Início: ' + (dataInicio || 'Não definida') +
                      '\\nData Fim: ' + (dataFim || 'Não definida') +
                      '\\n\\nEm ambiente real, os relatórios seriam filtrados por período.');
            }} else {{
                alert('Por favor, selecione pelo menos uma data para filtrar.');
            }}
        }}

        function limparFiltros() {{
            document.getElementById('data_inicio').value = '';
            document.getElementById('data_fim').value = '';
            alert('Filtros limpos!');
        }}

        // Definir data padrão para o último mês
        window.onload = function() {{
            const hoje = new Date();
            const umMesAtras = new Date();
            umMesAtras.setMonth(hoje.getMonth() - 1);

            document.getElementById('data_inicio').value = umMesAtras.toISOString().split('T')[0];
            document.getElementById('data_fim').value = hoje.toISOString().split('T')[0];
        }}
    </script>
</body>
</html>
'''

def generate_ocupacao_html(relatorio_ocupacao):
    """Gera HTML para relatório de ocupação"""
    html = ""
    for regiao, dados in relatorio_ocupacao.items():
        html += f"""
        <div class="relatorio-item">
            <h4>{regiao}</h4>
            <p><strong>Total:</strong> {sum(item['quantidade'] for item in dados)} placas</p>
        """
        for item in dados:
            html += f"""
            <p><strong>{item['Status_Atual']}:</strong> {item['quantidade']} placas
            (Faturamento: R$ {item['faturamento'] or 0:,.2f})</p>
            """
        html += "</div>"
    return html if html else "<p>Nenhum dado disponível</p>"

def generate_financeiro_html(relatorio_financeiro):
    """Gera HTML para relatório financeiro"""
    if not relatorio_financeiro:
        return "<p>Nenhum dado financeiro disponível</p>"

    html = f"""
    <div class="relatorio-item">
        <p><strong>Faturamento Total:</strong> R$ {relatorio_financeiro['faturamento_total']:,.2f}</p>
    </div>
    """

    html += """
    <div class="relatorio-item">
        <h4>Faturamento por Região</h4>
    """
    for item in relatorio_financeiro['faturamento_regiao']:
        html += f"""
        <p><strong>{item['Regiao']}:</strong> R$ {item['faturamento']:,.2f}</p>
        """
    html += "</div>"

    html += """
    <div class="relatorio-item">
        <h4>Faturamento por Tipo de Placa</h4>
    """
    for item in relatorio_financeiro['faturamento_tipo']:
        html += f"""
        <p><strong>{item['Tipo_Placa']}:</strong> R$ {item['faturamento']:,.2f}</p>
        """
    html += "</div>"

    html += """
    <div class="relatorio-item">
        <h4>Top 10 Clientes</h4>
    """
    for item in relatorio_financeiro['top_clientes']:
        html += f"""
        <p><strong>{item['Cliente_Locacao']}:</strong> R$ {item['faturamento']:,.2f}</p>
        """
    html += "</div>"

    return html

def generate_regiao_html(relatorio_regiao):
    """Gera HTML para relatório por região"""
    if not relatorio_regiao:
        return "<p>Nenhum dado por região disponível</p>"

    html = ""
    for item in relatorio_regiao:
        html += f"""
        <div class="relatorio-item">
            <h4>{item['Regiao']}</h4>
            <p><strong>Total de Placas:</strong> {item['total_placas']}</p>
            <p><strong>Disponíveis:</strong> {item['disponiveis']}</p>
            <p><strong>Locadas:</strong> {item['locadas']}</p>
            <p><strong>Reservadas:</strong> {item['reservadas']}</p>
            <p><strong>Manutenção:</strong> {item['manutencao']}</p>
            <p><strong>Faturamento Total:</strong> R$ {item['faturamento_total']:,.2f}</p>
        </div>
        """
    return html

# =============================================================================
# ROTAS PARA EXPORTAÇÃO DE RELATÓRIOS - COMPLETAMENTE CORRIGIDAS
# =============================================================================

@app.route('/relatorios/exportar/<tipo>')
def exportar_relatorio(tipo):
    """Exporta relatórios específicos - CORRIGIDA"""
    if 'user' not in session:
        return redirect('/login')

    try:
        if tipo == 'ocupacao':
            dados = relatorio_manager.gerar_relatorio_ocupacao()
            filename = f'relatorio_ocupacao_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return exportar_relatorio_ocupacao_excel(dados, filename)

        elif tipo == 'financeiro':
            dados = relatorio_manager.gerar_relatorio_financeiro()
            filename = f'relatorio_financeiro_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return exportar_relatorio_financeiro_excel(dados, filename)

        elif tipo == 'inventario':
            dados = relatorio_manager.gerar_relatorio_inventario()
            filename = f'relatorio_inventario_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            return exportar_relatorio_inventario_excel(dados, filename)

        else:
            return "Tipo de relatório inválido", 400

    except Exception as e:
        logging.error(f"❌ Erro ao exportar relatório: {e}")
        return f"Erro ao exportar relatório: {str(e)}", 500

def exportar_relatorio_ocupacao_excel(dados, filename):
    """Exporta relatório de ocupação para Excel - CORRIGIDA"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ocupação por Região"

        # Cabeçalhos
        headers = ["Região", "Status", "Quantidade", "Faturamento (R$)"]
        ws.append(headers)

        # Dados
        if dados:
            for regiao, items in dados.items():
                for item in items:
                    ws.append([
                        regiao,
                        item.get('Status_Atual', ''),
                        item.get('quantidade', 0),
                        item.get('faturamento', 0) or 0
                    ])

        # Formatar cabeçalhos - CORREÇÃO: Usar CORES_FORMATO_ARGB
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                   end_color=CORES_FORMATO_ARGB['primaria'],
                                   fill_type="solid")

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"❌ Erro ao exportar ocupação: {e}")
        return f"Erro ao exportar ocupação: {str(e)}", 500

def exportar_relatorio_financeiro_excel(dados, filename):
    """Exporta relatório financeiro para Excel - CORRIGIDA"""
    try:
        wb = Workbook()

        # Aba de faturamento por região
        if dados.get('faturamento_regiao'):
            ws1 = wb.active
            ws1.title = "Faturamento por Região"
            ws1.append(["Região", "Faturamento (R$)"])
            for item in dados['faturamento_regiao']:
                ws1.append([
                    item.get('Regiao', ''),
                    item.get('faturamento', 0) or 0
                ])

            # Formatar cabeçalhos - CORREÇÃO: Usar CORES_FORMATO_ARGB
            for col in range(1, 3):
                cell = ws1.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                       end_color=CORES_FORMATO_ARGB['primaria'],
                                       fill_type="solid")

        # Aba de faturamento por tipo
        if dados.get('faturamento_tipo'):
            ws2 = wb.create_sheet("Faturamento por Tipo")
            ws2.append(["Tipo de Placa", "Faturamento (R$)"])
            for item in dados['faturamento_tipo']:
                ws2.append([
                    item.get('Tipo_Placa', ''),
                    item.get('faturamento', 0) or 0
                ])

            # Formatar cabeçalhos - CORREÇÃO: Usar CORES_FORMATO_ARGB
            for col in range(1, 3):
                cell = ws2.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                       end_color=CORES_FORMATO_ARGB['primaria'],
                                       fill_type="solid")

        # Aba de top clientes
        if dados.get('top_clientes'):
            ws3 = wb.create_sheet("Top Clientes")
            ws3.append(["Cliente", "Faturamento (R$)"])
            for item in dados['top_clientes']:
                ws3.append([
                    item.get('Cliente_Locacao', ''),
                    item.get('faturamento', 0) or 0
                ])

            # Formatar cabeçalhos - CORREÇÃO: Usar CORES_FORMATO_ARGB
            for col in range(1, 3):
                cell = ws3.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                       end_color=CORES_FORMATO_ARGB['primaria'],
                                       fill_type="solid")

        # Ajustar largura das colunas em todas as abas
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"❌ Erro ao exportar financeiro: {e}")
        return f"Erro ao exportar financeiro: {str(e)}", 500

def exportar_relatorio_inventario_excel(dados, filename):
    """Exporta relatório de inventário para Excel - CORRIGIDA"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventário Completo"

        headers = ["Código", "Endereço", "Região", "Tipo", "Status", "Cliente", "Valor Mensal (R$)", "Data Cadastro"]
        ws.append(headers)

        if dados:
            for item in dados:
                ws.append([
                    item.get('Codigo_Ativo', ''),
                    item.get('Endereco', ''),
                    item.get('Regiao', ''),
                    item.get('Tipo_Placa', ''),
                    item.get('Status_Atual', ''),
                    item.get('Cliente_Locacao', '') or "Sem locação",
                    item.get('Valor_Mensal', 0),
                    item.get('Data_Cadastro', '')
                ])

        # Formatar cabeçalhos - CORREÇÃO: Usar CORES_FORMATO_ARGB
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                   end_color=CORES_FORMATO_ARGB['primaria'],
                                   fill_type="solid")

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logging.error(f"❌ Erro ao exportar inventário: {e}")
        return f"Erro ao exportar inventário: {str(e)}", 500

# =============================================================================
# MÓDULO DE EXPORTAÇÃO DE DADOS (COMPLETAMENTE CORRIGIDO)
# =============================================================================

@app.route('/exportar_dados')
def exportar_dados():
    """Página de exportação de dados - CORRIGIDA"""
    if 'user' not in session:
        return redirect('/login')

    return f'''
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Exportar Dados - SIG-ME</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: {CORES_FORMATO['claro']};
            color: {CORES_FORMATO['texto']};
        }}
        .header {{
            background: {CORES_FORMATO['branco']};
            padding: 20px 30px;
            box-shadow: 0 2px 20px rgba(0, 0, 0, 0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 4px solid {CORES_FORMATO['destaque']};
        }}
        .logo h1 {{
            color: {CORES_FORMATO['primaria']};
            font-size: 24px;
            font-weight: 700;
        }}
        .user-info {{
            display: flex;
            align-items: center;
            gap: 15px;
            color: {CORES_FORMATO['texto']};
            font-weight: 500;
        }}
        .btn-logout {{
            background: {CORES_FORMATO['destaque']};
            color: {CORES_FORMATO['branco']};
            padding: 10px 20px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
        }}
        .btn-logout:hover {{
            background: {CORES_FORMATO['alerta']};
            transform: translateY(-2px);
        }}
        .container {{
            max-width: 1400px;
            margin: 30px auto;
            padding: 0 25px;
        }}
        .navigation {{
            display: flex;
            gap: 12px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .nav-btn {{
            background: {CORES_FORMATO['primaria']};
            color: {CORES_FORMATO['branco']};
            padding: 14px 24px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .nav-btn:hover {{
            background: {CORES_FORMATO['secundaria']};
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(26, 54, 93, 0.2);
        }}
        .nav-btn.active {{
            background: {CORES_FORMATO['destaque']};
        }}
        .content {{
            background: {CORES_FORMATO['branco']};
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        }}
        .section-title {{
            font-size: 24px;
            font-weight: 700;
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid {CORES_FORMATO['claro']};
        }}
        .export-options {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-top: 30px;
        }}
        .export-card {{
            background: {CORES_FORMATO['claro']};
            padding: 25px;
            border-radius: 10px;
            text-align: center;
        }}
        .export-card h3 {{
            color: {CORES_FORMATO['primaria']};
            margin-bottom: 15px;
        }}
        .btn-exportar {{
            background: {CORES_FORMATO['sucesso']};
            color: {CORES_FORMATO['branco']};
            padding: 12px 24px;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            text-decoration: none;
            display: inline-block;
        }}
        .btn-exportar:hover {{
            background: {CORES_FORMATO['info']};
            transform: translateY(-2px);
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">
            <h1>🚀 SIG-ME - Exportar Dados</h1>
        </div>
        <div class="user-info">
            <span>👋 Olá, {session['user']['nome']} <small>({session['user']['perfil']})</small></span>
            <a href="/logout" class="btn-logout">🚪 Sair</a>
        </div>
    </div>

    <div class="container">
        <div class="navigation">
            <a href="/inventario" class="nav-btn">📋 Inventário</a>
            <a href="/relatorios" class="nav-btn">📊 Relatórios</a>
            <a href="/comercial" class="nav-btn">💼 Comercial</a>
            <a href="/contratos" class="nav-btn">📑 Contratos</a>
            <a href="/exportar_dados" class="nav-btn active">📥 Exportar Dados</a>
        </div>

        <div class="content">
            <h2 class="section-title">📥 Exportar Dados</h2>

            <div class="export-options">
                <div class="export-card">
                    <h3>📋 Exportar Inventário</h3>
                    <p>Exportar todas as placas do inventário em formato Excel</p>
                    <a href="/exportar/inventario" class="btn-exportar">📥 Exportar Excel</a>
                </div>

                <div class="export-card">
                    <h3>👥 Exportar Clientes</h3>
                    <p>Exportar lista de clientes em formato Excel</p>
                    <a href="/exportar/clientes" class="btn-exportar">📥 Exportar Excel</a>
                </div>

                <div class="export-card">
                    <h3>📑 Exportar Contratos</h3>
                    <p>Exportar lista de contratos em formato Excel</p>
                    <a href="/exportar/contratos" class="btn-exportar">📥 Exportar Excel</a>
                </div>

                <div class="export-card">
                    <h3>📊 Exportar Relatórios</h3>
                    <p>Exportar relatórios completos em Excel</p>
                    <a href="/relatorios" class="btn-exportar">📊 Ver Relatórios</a>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
'''

@app.route('/exportar/inventario')
def exportar_inventario():
    """Exporta o inventário para Excel - CORRIGIDA"""
    if 'user' not in session:
        return redirect('/login')

    try:
        db = DatabaseConnection()
        placas = db.execute_query("""
            SELECT Codigo_Ativo, Endereco, Regiao, Tipo_Placa, Status_Atual,
                   Cliente_Locacao, Valor_Mensal, Data_Cadastro
            FROM placas
            ORDER BY Regiao, Codigo_Ativo
        """) or []

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventário"

        # Cabeçalhos
        headers = ["Código", "Endereço", "Região", "Tipo", "Status", "Cliente", "Valor Mensal", "Data Cadastro"]
        ws.append(headers)

        # Adicionar dados
        for placa in placas:
            ws.append([
                placa['Codigo_Ativo'],
                placa['Endereco'],
                placa['Regiao'],
                placa['Tipo_Placa'],
                placa['Status_Atual'],
                placa['Cliente_Locacao'] or "Sem locação",
                float(placa['Valor_Mensal']) if placa['Valor_Mensal'] else 0,
                placa['Data_Cadastro']
            ])

        # Formatação - CORREÇÃO: Usar CORES_FORMATO_ARGB
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                   end_color=CORES_FORMATO_ARGB['primaria'],
                                   fill_type="solid")

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em um buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'inventario_sigme_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"❌ Erro ao exportar inventário: {e}")
        return f"Erro ao exportar inventário: {str(e)}", 500

@app.route('/exportar/clientes')
def exportar_clientes():
    """Exporta a lista de clientes para Excel - CORRIGIDA"""
    if 'user' not in session:
        return redirect('/login')

    try:
        db = DatabaseConnection()
        clientes = db.execute_query("""
            SELECT Nome_Fantasia, Nome_Razao_Social, CNPJ_CPF, Telefone, Email, created_at
            FROM clientes
            ORDER BY Nome_Fantasia
        """) or []

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Cabeçalhos
        headers = ["Nome Fantasia", "Razão Social", "CNPJ/CPF", "Telefone", "Email", "Data Cadastro"]
        ws.append(headers)

        # Adicionar dados
        for cliente in clientes:
            ws.append([
                cliente['Nome_Fantasia'],
                cliente['Nome_Razao_Social'],
                cliente['CNPJ_CPF'],
                cliente['Telefone'],
                cliente['Email'],
                cliente['created_at']
            ])

        # Formatação - CORREÇÃO: Usar CORES_FORMATO_ARGB
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                   end_color=CORES_FORMATO_ARGB['primaria'],
                                   fill_type="solid")

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em um buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'clientes_sigme_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"❌ Erro ao exportar clientes: {e}")
        return f"Erro ao exportar clientes: {str(e)}", 500

@app.route('/exportar/contratos')
def exportar_contratos():
    """Exporta a lista de contratos para Excel - CORRIGIDA"""
    if 'user' not in session:
        return redirect('/login')

    try:
        db = DatabaseConnection()
        contratos = db.execute_query("""
            SELECT c.*, cl.Nome_Fantasia as cliente_nome, f.razao_social as fornecedor_nome
            FROM contratos c
            LEFT JOIN clientes cl ON c.cliente_id = cl.id
            LEFT JOIN fornecedores f ON c.fornecedor_id = f.id
            ORDER BY c.data_inicio DESC
        """) or []

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Contratos"

        # Cabeçalhos
        headers = ["ID", "Tipo", "Cliente/Fornecedor", "Descrição", "Valor Mensal",
                   "Data Início", "Data Fim", "Status", "Arquivo PDF", "Data Cadastro"]
        ws.append(headers)

        # Adicionar dados
        for contrato in contratos:
            # Determinar nome do contratante
            if contrato['tipo'] == 'cliente':
                nome_contratante = contrato.get('cliente_nome', 'N/A')
            else:
                nome_contratante = contrato.get('fornecedor_nome', 'N/A')

            ws.append([
                contrato['id'],
                contrato['tipo'],
                nome_contratante,
                contrato['descricao'],
                float(contrato['valor_mensal']) if contrato['valor_mensal'] else 0,
                contrato['data_inicio'],
                contrato['data_fim'] or 'N/A',
                contrato['status'],
                contrato['arquivo_pdf'] or 'N/A',
                contrato['created_at']
            ])

        # Formatação - CORREÇÃO: Usar CORES_FORMATO_ARGB
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=CORES_FORMATO_ARGB['primaria'],
                                   end_color=CORES_FORMATO_ARGB['primaria'],
                                   fill_type="solid")

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Salvar em um buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'contratos_sigme_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"❌ Erro ao exportar contratos: {e}")
        return f"Erro ao exportar contratos: {str(e)}", 500

# =============================================================================
# INICIALIZAÇÃO DO SISTEMA
# =============================================================================
if __name__ == '__main__':
    if inicializar_database():
        app.run(debug=DEBUG, host='0.0.0.0', port=5000)
    else:
        logging.error("Falha ao inicializar o sistema!")
